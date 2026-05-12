"""
slides_agent.py  —  ADK Automated Slides Report Tool  (Phase 2)
================================================================
Transforms a weekly report Excel file into a polished Google Slides
presentation by:

  1. Extracting text metadata, executive summary, and task data from Excel
  2. Dynamically detecting the Timeline sheet range (no fixed row/col assumptions)
  3. Rendering the Timeline range to a high-resolution PNG via openpyxl + Pillow
  4. Uploading the PNG to Google Drive
  5. Duplicating the template Google Slides presentation
  6. Injecting all {{PLACEHOLDER}} values via Slides API find-and-replace
  7. Replacing {{TIMELINE_SCREENSHOT}} shape with the PNG, smart-scaled to fit
  8. Deleting the temporary Drive image
  9. Returning the final Slides URL to the user

Architecture (SequentialAgent):
  1. ExcelReaderAgent   → reads Excel → session["excel_data"], session["timeline_range"]
  2. ScreenshotAgent    → renders timeline range → session["drive_image_id"], session["image_dimensions"]
  3. SlidesBuilderAgent → duplicates template, injects text + image → session["slides_url"]
  4. CleanupAgent       → deletes temp Drive image, returns final URL

Configuration (via .env):
  GOOGLE_SLIDES_TEMPLATE_ID      — ID of the master template presentation
  GOOGLE_OAUTH_CLIENT_SECRET_JSON — path to OAuth2 client secret JSON
  GOOGLE_OAUTH_TOKEN_JSON        — path to cached OAuth token (auto-created)
  GOOGLE_SLIDES_PARENT_FOLDER_ID — (optional) Drive folder for output
  GEMINI_MODEL                   — default: gemini-2.5-flash
"""

import io
import json
import math
import os
import re
import secrets
import shutil
import subprocess
import tempfile
from datetime import datetime, timedelta
from typing import Optional
from urllib.parse import unquote, urlparse

# ── openpyxl / Pillow (Excel → PNG) ──────────────────────────────────────────
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill, GradientFill
from PIL import Image, ImageDraw, ImageFont

# ── Matplotlib (Gantt Chart) ────────────────────────────────────────────────
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime

def _format_indian_date(date_str: str) -> str:
    """Convert date to Indian format DD/MM/YYYY"""
    if not date_str:
        return ""
    date_str = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y"):
        try:
            dt = datetime.strptime(date_str.split()[0], fmt)
            return dt.strftime("%d/%m/%Y")
        except:
            continue
    return date_str

# ── Google Auth ───────────────────────────────────────────────────────────────
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload
from googleapiclient.errors import HttpError
from google.adk.agents import LlmAgent, SequentialAgent
from google.adk.artifacts.artifact_util import parse_artifact_uri
from google.adk.tools import ToolContext
from google.genai import types as genai_types
from google.genai import Client as GenaiClient

# ══════════════════════════════════════════════════════════════════════════════
# ENVIRONMENT
# ══════════════════════════════════════════════════════════════════════════════

load_dotenv()

TEMPLATE_ID          = os.environ.get("GOOGLE_SLIDES_TEMPLATE_ID", "").strip()
TIMELINE_SHEET_NAME  = os.environ.get("TIMELINE_SHEET_NAME",  "Timeline 1").strip()
OAUTH_SECRET_PATH    = os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET_JSON", "").strip()
OAUTH_TOKEN_PATH     = os.environ.get("GOOGLE_OAUTH_TOKEN_JSON",
                                       "./google_oauth_token.json").strip()
OAUTH_AUTH_MODE      = os.environ.get("GOOGLE_OAUTH_AUTH_MODE", "local_server").strip()
PARENT_FOLDER_ID     = os.environ.get("GOOGLE_SLIDES_PARENT_FOLDER_ID", "").strip()
OUTPUT_DIR           = os.environ.get("OUTPUT_DIR", "./outputs").strip()
APP_ROOT             = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
GEMINI_MODEL         = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash").strip()
RETRY_ATTEMPTS       = int(os.environ.get("RETRY_ATTEMPTS",    "5"))
RETRY_DELAY          = float(os.environ.get("RETRY_INITIAL_DELAY", "2"))

def _get_genai_client() -> GenaiClient:
    api_key = (
        os.environ.get("GEMINI_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or os.environ.get("GOOGLE_GENAI_API_KEY")
    )
    if not api_key:
        raise ValueError(
            "Missing Gemini API key. Set GEMINI_API_KEY or GOOGLE_API_KEY."
        )
    return GenaiClient(api_key=api_key)


def _generate_executive_summary_llm(completed_tasks, in_progress_tasks, upcoming_tasks) -> str:
    """Use LLM to generate a 60-word narrative executive summary."""
    if not completed_tasks and not in_progress_tasks and not upcoming_tasks:
        return "This week involved planning and coordination activities. Team remains aligned on project objectives."
    
    # Build task summary for LLM
    completed = ", ".join(completed_tasks[:3]) if completed_tasks else "none"
    in_progress = ", ".join(in_progress_tasks[:3]) if in_progress_tasks else "none"
    upcoming = ", ".join(upcoming_tasks[:3]) if upcoming_tasks else "none"
    
    prompt = f"""Write a 60-word executive summary for a weekly project status report.

Completed tasks: {completed}
In progress: {in_progress}
Upcoming: {upcoming}

Requirements:
- Write in first person plural ("we focused", "our team")
- No numbers, percentages, or stats
- Sound natural, like a story
- Focus on what was accomplished and what's next
- Exactly 60 words"""

    try:
        client = _get_genai_client()
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=_RETRY_CONFIG,
        )
        if response.text:
            return response.text.strip()
    except Exception as e:
        print(f"[DEBUG] LLM summary generation error: {e}")
    
    # Fallback if LLM fails
    if completed_tasks:
        return f"This week we focused on {completed_tasks[0].lower()}. Moving forward, our efforts center on {in_progress_tasks[0].lower() if in_progress_tasks else 'ongoing development'}."
    return "Project progress continues with active development and planning."


def _populate_slide_fields_llm(placeholder_map, task_data_rows) -> dict:
    """Use LLM to populate slide fields based on task data."""
    
    # Build task context for LLM
    tasks_text = ""
    if task_data_rows and len(task_data_rows) > 1:
        header = [str(h).strip().lower() for h in task_data_rows[0]]
        task_idx = next((i for i, h in enumerate(header) if "task" in h), None)
        module_idx = next((i for i, h in enumerate(header) if "module" in h), None)
        status_idx = next((i for i, h in enumerate(header) if "status" in h), None)
        for row in task_data_rows[1:8]:  # First 7 tasks
            if not row:
                continue
            task_val = ""
            if task_idx is not None and len(row) > task_idx and row[task_idx]:
                task_val = str(row[task_idx]).strip()
            if not task_val and row:
                task_val = str(row[0]).strip()
            if not task_val:
                continue
            module_val = ""
            if module_idx is not None and len(row) > module_idx and row[module_idx]:
                module_val = str(row[module_idx]).strip()
            status_val = ""
            if status_idx is not None and len(row) > status_idx and row[status_idx]:
                status_val = str(row[status_idx]).strip()
            elif row and row[-1]:
                status_val = str(row[-1]).strip()
            if module_val:
                tasks_text += f"- {task_val} | Module: {module_val} | Status: {status_val or 'Unknown'}\n"
            else:
                tasks_text += f"- {task_val} | Status: {status_val or 'Unknown'}\n"
    
    # ── IMPROVED PROMPT ────────────────────────────────────────────────────
    prompt = f"""You are creating professional content for a Google Slides weekly status report.

Task data from spreadsheet:
{tasks_text}

CRITICAL FORMATTING RULES — follow exactly:

1. EXECUTIVE SUMMARY (60 words, narrative paragraph, first-person plural, no bullets, no numbers)

2. KEY ACTIVITIES COMPLETED / IN PROGRESS — use this EXACT two-level structure:
   • Bold module/category name (e.g. "BigQuery Data Architecture")
     ○ One concise action sentence describing what was done
     ○ Another concise action sentence if needed
   • Bold next module/category name
     ○ One concise action sentence

3. KEY ACTIVITIES UPCOMING — same two-level structure:
   • Bold deliverable/category name (e.g. "Business Logic Alignment")
     ○ One concise action sentence describing what will be done

FORMAT RULES:
- Level 1 (•): Bold category/module name only — NO action verbs, just the topic label
- Level 2 (○): Concise complete sentence describing the specific activity (max 15 words)
- Group related completed AND in-progress tasks together under one bold heading
- NEVER use semicolons to chain multiple actions in one bullet
- NEVER use long run-on sentences with ";" separating tasks
- Each ○ sub-bullet must be a standalone, complete, short sentence
- Max 3 level-1 bullets per section, max 2 sub-bullets per level-1

EXAMPLE OUTPUT for "Key Activities Completed":
• BigQuery Data Architecture
  ○ Successfully created and validated views for all data sources within BigQuery.
• NLP-to-SQL Development
  ○ Developed the initial version of the NLP-to-SQL output agent.
  ○ Integrated a Looker dashboard to visualize agent-generated insights.
• Advanced Agent Features (In Progress)
  ○ Developing a data mining agent capable of performing automated web searches.
  ○ Integrating Google Maps API into the NLP agent for geospatial intelligence.

Return ONLY valid JSON with these exact keys:
- "Executive Summary" (60-word narrative paragraph, no bullets)
- "Key Activities Completed" (two-level bullet structure as shown above)
- "Key Activities In Progress" (two-level bullet structure — only if tasks differ from completed section)
- "Key Activities Upcoming" (two-level bullet structure as shown above)

JSON format:
{{
  "Executive Summary": "This week we focused on...",
  "Key Activities Completed": "• Module Name\\n  ○ Concise sentence.\\n• Another Module\\n  ○ Concise sentence.",
  "Key Activities In Progress": "• Module Name\\n  ○ Concise sentence.",
  "Key Activities Upcoming": "• Deliverable Name\\n  ○ Concise sentence."
}}

Return ONLY valid JSON, no extra text or markdown backticks:"""

    try:
        client = _get_genai_client()
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=_RETRY_CONFIG,
        )
        
        # Parse JSON from response
        text = response.text.strip()
        if "{" in text and "}" in text:
            start = text.find("{")
            end = text.rfind("}") + 1
            json_str = text[start:end]
            result = json.loads(json_str)
            print(f"[DEBUG] LLM field population success: {list(result.keys())}")
            return result
    except Exception as e:
        print(f"[DEBUG] LLM field population error: {e}")
    
    return {}


def _fill_slide2_via_llm(slides_svc, presentation_id: str, placeholder_map: dict, task_data: list):
    """
    Fetch all shapes on slide 2, use LLM to determine what content belongs
    in each shape based on its existing text/placeholder hint, then directly
    overwrite the text. This bypasses find-and-replace entirely.
    """
    try:
        prs = slides_svc.presentations().get(presentationId=presentation_id).execute()
        all_slides = prs.get("slides", [])
        
        if len(all_slides) < 2:
            print("[DEBUG _fill_slide2] Less than 2 slides found")
            return
        
        slide2 = all_slides[1]  # 0-indexed → slide 2
        
        # Collect all text shapes on slide 2 with their current text
        shapes_info = []
        for elem in slide2.get("pageElements", []):
            obj_id = elem.get("objectId", "")
            title  = elem.get("title", "")
            shape  = elem.get("shape", {})
            text_obj = shape.get("text", {})
            
            full_text = ""
            for te in text_obj.get("textElements", []):
                content = te.get("textRun", {}).get("content", "")
                if content:
                    full_text += content
            
            full_text = full_text.strip()
            if full_text or title:
                shapes_info.append({
                    "objectId": obj_id,
                    "title":    title,
                    "text":     full_text,
                })
        
        print(f"[DEBUG _fill_slide2] Found {len(shapes_info)} text shapes on slide 2")
        for s in shapes_info:
            print(f"  → id={s['objectId']}, title='{s['title']}', text='{s['text'][:60]}'")
        
        if not shapes_info:
            print("[DEBUG _fill_slide2] No shapes found on slide 2")
            return
        
        # Build content context for LLM
        exec_summary    = placeholder_map.get("{{Executive Summary}}", "")
        completed       = placeholder_map.get("{{Key Activities Completed}}", "")
        in_progress     = placeholder_map.get("{{Key Activities In Progress}}", "")
        upcoming        = placeholder_map.get("{{Key Activities Upcoming}}", "")
        project_name    = placeholder_map.get("{{Project Name}}", "")
        from_date       = placeholder_map.get("{{From Date}}", "")
        to_date         = placeholder_map.get("{{To Date}}", "")
        
        # Build tasks summary for LLM context
        task_summary = ""
        if task_data and len(task_data) > 1:
            for row in task_data[1:8]:
                if row and len(row) >= 4:
                    task_summary += f"- {row[0]} | Status: {row[-1]}\n"
        
        shapes_json = json.dumps([
            {"objectId": s["objectId"], "title": s["title"], "currentText": s["text"]}
            for s in shapes_info
        ], indent=2)
        
        # ── IMPROVED PROMPT ────────────────────────────────────────────────
        prompt = f"""You are filling a professional weekly status report slide (Slide 2).

Project: {project_name}
Period: {from_date} to {to_date}

Available content:
- Executive Summary (60 words): {exec_summary}
- Completed Tasks: {completed}
- In Progress Tasks: {in_progress}
- Upcoming/To-Do Tasks: {upcoming}

These are the text shapes currently on Slide 2 of the Google Slides presentation:
{shapes_json}

Your task: For each shape, decide what content to place in it based on its "currentText" placeholder.

CRITICAL FORMATTING RULES:

Executive Summary shape:
- Exactly 60 words, first person plural ("we", "our team"), narrative prose
- NO bullet points, NO bold markers, NO semicolons chaining tasks

Activities shapes — preserve ORIGINAL task descriptions (DO NOT truncate):

Completed / In Progress activities shape — use this structure:
Category Name (from original task, use exact wording)
Full original task description (preserve ALL words, do not abbreviate)
Next Category
Full original task description

Upcoming / To-Do activities shape — same structure:
Deliverable Name
Full original task description

FORMAT RULES:
- Each line is one item (no manual bullets or dashes needed)
- First line of each item: category/module label (will be bolded via API)
- Following lines: Include the FULL original task description (no word limit)
- NEVER truncate or shorten task descriptions
- Include all tasks (one per paragraph, separated by newlines)
- DO NOT include bullet characters (•) or dashes - API will add them
- If a shape is a title/header (short text like "Weekly Status"), keep original or use project name
- If a shape doesn't match any category, set "content" to null

EXAMPLE of correct format (preserve full task descriptions - NO bullets, API will add them):
Requirements & Discovery Workshops
Conduct Requirements & Discovery Workshops; Finalize Architecture & Solution Design.
Web Portal & File Validation
Build a secure web portal for uploading/validating engineering data files/templates (Excel/CSV); Implement automated schema detection and file validation.
Orchestration Layer
Create intelligent orchestration layer; Interpret embedded instructions; Extract customer identifiers; Process default values, cross-column dependencies, formula-based computations...

Return ONLY a valid JSON array, no extra text, no markdown:
[
  {{"objectId": "...", "content": "text to place here or null"}},
  ...
]"""

        try:
            client = _get_genai_client()
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=prompt,
                config=_RETRY_CONFIG,
            )
            text = response.text.strip()
            
            # Parse JSON
            if "[" in text and "]" in text:
                start = text.find("[")
                end   = text.rfind("]") + 1
                assignments = json.loads(text[start:end])
            else:
                print("[DEBUG _fill_slide2] LLM returned no valid JSON array")
                return
            
            print(f"[DEBUG _fill_slide2] LLM assigned {len(assignments)} shapes")
        except Exception as e:
            print(f"[DEBUG _fill_slide2] LLM call failed: {e}")
            return
        
        # Build batchUpdate requests to replace text in each shape
        requests = []
        shape_lookup = {}
        for slide in all_slides:
            for elem in slide.get("pageElements", []):
                obj_id = elem.get("objectId", "")
                if obj_id:
                    shape_lookup[obj_id] = elem.get("shape", {})
        
        for assignment in assignments:
            obj_id = assignment.get("objectId", "")
            content = assignment.get("content")
            
            if not obj_id or content is None:
                continue
            
            content = str(content).strip()
            if not content:
                continue
            
            # Get original text to check if this is a template header
            original_text = ""
            for s in shapes_info:
                if s["objectId"] == obj_id:
                    original_text = s.get("text", "").strip()
                    break
            
            # If original text is all caps (template header), DON'T modify at all
            if original_text.isupper() and len(original_text) < 60:
                print(f"[DEBUG] {obj_id}: SKIP header - keep original: '{original_text}'")
                continue
            
            # NEW: If content is all caps (like a header), use original text instead
            if content.isupper() and len(content) < 60:
                print(f"[DEBUG] {obj_id}: SKIP new header - keep original: '{original_text}'")
                continue
            
            # Check if shape has existing text before deleting
            shape = shape_lookup.get(obj_id, {})
            has_text = any(
                te.get("textRun", {}).get("content", "").strip()
                for te in shape.get("text", {}).get("textElements", [])
            )
            
            if has_text:
                requests.append({
                    "deleteText": {
                        "objectId": obj_id,
                        "textRange": {"type": "ALL"}
                    }
                })
            requests.append({
                "insertText": {
                    "objectId":       obj_id,
                    "insertionIndex": 0,
                    "text":           content,
                }
            })
            requests.append({
                "insertText": {
                    "objectId":       obj_id,
                    "insertionIndex": 0,
                    "text":           content,
                }
            })
            
            template_headers = ["executive summary", "key activities", "completed", "in progress", "upcoming", "period"]
            content_lower = content.lower()
            is_template_header = any(h in content_lower for h in template_headers) and len(content) < 40
            
            has_multiple_lines = content.count("\n") >= 2
            
            # Always delete existing text first
            if has_text:
                requests.append({
                    "deleteText": {
                        "objectId": obj_id,
                        "textRange": {"type": "ALL"}
                    }
                })
            
            # Insert new content
            requests.append({
                "insertText": {
                    "objectId":       obj_id,
                    "insertionIndex": 0,
                    "text":           content,
                }
            })
            
            if is_template_header:
                print(f"[DEBUG] {obj_id}: SKIP styling (template header): '{content[:30]}...'")
            else:
                # Apply base style (DM Sans, consistent with other slides)
                requests.append({
                    "updateTextStyle": {
                        "objectId": obj_id,
                        "textRange": {"type": "ALL"},
                        "style": {
                            "fontFamily": "DM Sans",
                            "fontSize": {"magnitude": 11, "unit": "PT"},
                        },
                        "fields": "fontFamily,fontSize"
                    }
                })
                
                if has_multiple_lines:
                    para_requests = _build_bold_category_headers(obj_id, content)
                    requests.extend(para_requests)
                    print(f"[DEBUG] {obj_id}: Applied base style + {len(para_requests)} bold ranges for category headers")
                else:
                    print(f"[DEBUG] {obj_id}: Applied base style DM Sans 11pt")
        
        if requests:
            slides_svc.presentations().batchUpdate(
                presentationId=presentation_id,
                body={"requests": requests}
            ).execute()
            print(f"[DEBUG _fill_slide2] Applied {len(requests)} text updates to slide 2")
        else:
            print("[DEBUG _fill_slide2] No requests to apply")
    
    except Exception as e:
        import traceback
        print(f"[DEBUG _fill_slide2] Error (non-fatal): {e}")
        traceback.print_exc()


# Slide canvas dimensions (inches → used for smart-scaling)
SLIDE_W_IN = 10.0   # standard widescreen width
SLIDE_H_IN = 5.63   # standard widescreen height
SLIDE_W_PX = int(SLIDE_W_IN * 96)   # 960 px  (96 dpi reference)
SLIDE_H_PX = int(SLIDE_H_IN * 96)   # 540 px

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/presentations",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]

_RETRY_CONFIG = genai_types.GenerateContentConfig(
    http_options=genai_types.HttpOptions(
        retry_options=genai_types.HttpRetryOptions(
            initial_delay=RETRY_DELAY,
            attempts=RETRY_ATTEMPTS,
        ),
    ),
)

def _build_bold_category_headers(obj_id: str, content: str) -> list:
    """Bold category headers in activity content - bold lines 0, 2, 4, (the category names)."""
    requests = []
    lines = content.split("\n")
    current_pos = 0
    
    for i, line in enumerate(lines):
        if i == 0:
            line_start = 0
        else:
            line_start = current_pos
        
        line_length = len(line)
        
        if line_length > 0 and i % 2 == 0:
            requests.append({
                "updateTextStyle": {
                    "objectId": obj_id,
                    "textRange": {
                        "startIndex": line_start,
                        "endIndex": line_start + line_length
                    },
                    "style": {"bold": True},
                    "fields": "bold"
                }
            })
        
        current_pos = line_start + line_length + 1
    
    return requests

# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE AUTH HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _get_credentials() -> Credentials:
    """Load or refresh OAuth2 user credentials."""
    secret = os.path.abspath(os.path.expanduser(OAUTH_SECRET_PATH))
    token  = os.path.abspath(os.path.expanduser(OAUTH_TOKEN_PATH))

    if not os.path.exists(secret):
        raise FileNotFoundError(
            f"OAuth client secret not found: {secret}\n"
            "Set GOOGLE_OAUTH_CLIENT_SECRET_JSON in your .env file."
        )

    creds = None
    if os.path.exists(token):
        creds = Credentials.from_authorized_user_file(token, GOOGLE_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(secret, GOOGLE_SCOPES)
            if OAUTH_AUTH_MODE == "console":
                creds = flow.run_console()
            else:
                try:
                    creds = flow.run_local_server(port=0)
                except Exception:
                    creds = flow.run_console()

        os.makedirs(os.path.dirname(token) or ".", exist_ok=True)
        with open(token, "w") as f:
            f.write(creds.to_json())

    return creds


def _build_services():
    creds  = _get_credentials()
    drive  = build("drive",         "v3", credentials=creds)
    slides = build("slides",        "v1", credentials=creds)
    return drive, slides


# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _safe_json(value):
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value) if value else None
    except Exception:
        return None


def _slugify(text: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_\-]", "_", (text or "").strip())[:50]


def _na(value) -> str:
    """Return value as string or 'N/A' if empty."""
    s = str(value).strip() if value is not None else ""
    return s if s else "N/A"


def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
    """Convert 6 or 8-char hex string to (R, G, B)."""
    h = hex_color.lstrip("#").upper()
    if len(h) == 8:   # ARGB format used by openpyxl
        h = h[2:]     # strip alpha
    if len(h) != 6:
        return (255, 255, 255)
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))


def _cell_bg_rgb(cell) -> Optional[tuple[int, int, int]]:
    """Extract cell background colour as (R,G,B) or None if no fill."""
    try:
        fill = cell.fill
        if isinstance(fill, PatternFill) and fill.fill_type == "solid":
            fg = fill.fgColor
            if fg.type == "rgb" and fg.rgb and fg.rgb.upper() not in ("00000000", "FFFFFFFF"):
                return _hex_to_rgb(fg.rgb)
        if isinstance(fill, GradientFill):
            if fill.stop:
                rgb = fill.stop[0].color.rgb
                return _hex_to_rgb(rgb)
    except Exception:
        pass
    return None


_EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
_EXCEL_IMAGE_MIME_TYPES = {
    "image/png",
    "image/jpeg",
    "image/gif",
    "image/bmp",
    "image/tiff",
}
_EXCEL_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")


def _strip_unsupported_excel_attachments(callback_context, llm_request):
    """
    Remove Excel file parts from LLM requests to avoid unsupported MIME errors.
    Replaces them with a text hint containing the file URI or display name.
    """
    if not llm_request or not getattr(llm_request, "contents", None):
        return None

    for content in llm_request.contents:
        parts = getattr(content, "parts", None) or []
        if not parts:
            continue

        new_parts = []
        for part in parts:
            file_data = getattr(part, "file_data", None)
            inline_data = getattr(part, "inline_data", None)
            
            # Skip image data that might be embedded in Excel uploads
            if inline_data:
                mime = (getattr(inline_data, "mime_type", None) or "").lower()
                if mime in _EXCEL_IMAGE_MIME_TYPES:
                    new_parts.append(genai_types.Part(
                        text="[Image attachment removed - not supported]"
                    ))
                    continue

            if file_data:
                mime = (file_data.mime_type or "").lower()
                if mime in _EXCEL_MIME_TYPES:
                    file_ref = file_data.file_uri or file_data.display_name or "uploaded Excel file"
                    stored_path = ""
                    if file_data.file_uri:
                        callback_context.state["uploaded_excel_uri"] = file_data.file_uri
                        parsed = parse_artifact_uri(file_data.file_uri)
                        if parsed:
                            callback_context.state["uploaded_excel_artifact_name"] = parsed.filename
                            callback_context.state["uploaded_excel_artifact_version"] = parsed.version
                        if file_data.file_uri.startswith("file://"):
                            local_path = _normalize_excel_path(file_data.file_uri)
                            if os.path.exists(local_path):
                                callback_context.state["uploaded_excel_path"] = local_path
                                stored_path = local_path
                    if file_data.display_name:
                        callback_context.state["uploaded_excel_name"] = file_data.display_name
                    new_parts.append(genai_types.Part(
                        text=f"Uploaded Excel file path/URI: {stored_path or file_ref}"
                    ))
                    continue

            if inline_data:
                mime = (inline_data.mime_type or "").lower()
                if mime in _EXCEL_MIME_TYPES:
                    display = getattr(inline_data, "display_name", None) or "uploaded Excel file"
                    data = getattr(inline_data, "data", None)
                    saved_path = ""
                    if data:
                        saved_path = _write_excel_bytes(data, display)
                        callback_context.state["uploaded_excel_path"] = saved_path
                        callback_context.state["uploaded_excel_name"] = display
                    new_parts.append(genai_types.Part(
                        text=f"Uploaded Excel file path: {saved_path or display}"
                    ))
                    continue

            new_parts.append(part)

        content.parts = new_parts

    return None


_EXCEL_PATH_RE = re.compile(
    r"((?:file://)?(?:/[^\s\"']+|[A-Za-z]:[\\/][^\s\"']+)\.(?:xlsx|xlsm|xltx|xltm|xls))",
    re.IGNORECASE,
)


def _extract_excel_path(value: str) -> str:
    """Extract an Excel path/URI from a user-provided string."""
    s = (value or "").strip().strip("\"'")
    m = _EXCEL_PATH_RE.search(s)
    if m:
        return m.group(1).rstrip(").,;")
    return s.rstrip(").,;")


def _normalize_excel_path(value: str) -> str:
    """Normalize local paths and file:// URIs into a local filesystem path."""
    raw = _extract_excel_path(value)
    if raw.lower().startswith("file://"):
        parsed = urlparse(raw)
        raw = unquote(parsed.path)
        # Windows file URI: file:///C:/path/to/file.xlsx
        if re.match(r"^/[A-Za-z]:/", raw):
            raw = raw[1:]

    raw = raw.strip().strip("\"'")
    return os.path.abspath(os.path.expanduser(raw))


def _download_google_sheet(sheets_url: str) -> tuple[Optional[str], str]:
    """Download a Google Sheet as Excel and return local path."""
    import traceback
    # Extract spreadsheet ID from URL
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', sheets_url)
    if not match:
        return None, "Invalid Google Sheets URL"
    
    spreadsheet_id = match.group(1)
    print(f"[DEBUG] Google Sheet ID: {spreadsheet_id}")
    
    try:
        # Build Drive service
        creds = _get_credentials()
        drive = build("drive", "v3", credentials=creds)
        
        print(f"[DEBUG] Attempting to export Google Sheet...")
        
        # Export as Excel
        response = drive.files().export(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        print(f"[DEBUG] Export response status: {response.status}")
        
        # Save to local file
        output_dir = _get_output_dir()
        out_path = os.path.join(output_dir, f"google_sheet_{spreadsheet_id}.xlsx")
        
        with open(out_path, "wb") as f:
            f.write(response.content)
        
        print(f"[DEBUG] Downloaded Google Sheet to: {out_path}")
        return out_path, ""
        
    except Exception as e:
        print(f"[ERROR] Google Sheet download failed: {e}")
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return None, f"Failed to download Google Sheet: {e}"


def _get_output_dir() -> str:
    output_dir = os.path.expanduser(OUTPUT_DIR)
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(APP_ROOT, output_dir)
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def _safe_excel_filename(name: str) -> str:
    base = os.path.basename((name or "").strip()) or "uploaded_excel.xlsx"
    root, ext = os.path.splitext(base)
    if ext.lower() not in _EXCEL_EXTS:
        ext = ".xlsx"
    return f"{root or 'uploaded_excel'}{ext}"


def _write_excel_bytes(data: bytes, display_name: str) -> str:
    out_dir = _get_output_dir()
    base = _safe_excel_filename(display_name)
    path = os.path.join(out_dir, base)
    if os.path.exists(path):
        root, ext = os.path.splitext(base)
        path = os.path.join(out_dir, f"{root}_{secrets.token_hex(4)}{ext}")
    with open(path, "wb") as f:
        f.write(data)
    return path


async def _materialize_excel_artifact(
    tool_context: ToolContext,
    artifact_name: str,
    version: Optional[int] = None,
    preferred_name: str = "",
) -> str:
    try:
        part = await tool_context.load_artifact(filename=artifact_name, version=version)
    except Exception:
        return ""
    if not part:
        return ""

    inline_data = getattr(part, "inline_data", None)
    if inline_data and getattr(inline_data, "data", None):
        display = preferred_name or getattr(inline_data, "display_name", None) or artifact_name
        return _write_excel_bytes(inline_data.data, display)

    file_data = getattr(part, "file_data", None)
    if file_data and file_data.file_uri and file_data.file_uri.startswith("file://"):
        local_path = _normalize_excel_path(file_data.file_uri)
        if os.path.exists(local_path):
            return local_path
    return ""


async def _resolve_excel_from_state(
    tool_context: ToolContext,
    provided_path: str,
) -> str:
    state = tool_context.state

    for key in ("uploaded_excel_path", "excel_file_path"):
        candidate = state.get(key, "")
        if candidate:
            resolved = _resolve_excel_path(_normalize_excel_path(candidate))
            if os.path.exists(resolved):
                return resolved

    artifact_name = state.get("uploaded_excel_artifact_name", "")
    artifact_version = state.get("uploaded_excel_artifact_version")
    if artifact_name:
        path = await _materialize_excel_artifact(
            tool_context,
            artifact_name,
            version=artifact_version,
            preferred_name=state.get("uploaded_excel_name", ""),
        )
        if path:
            state["uploaded_excel_path"] = path
            return path

    uploaded_uri = state.get("uploaded_excel_uri", "")
    if uploaded_uri:
        if uploaded_uri.startswith("artifact://"):
            parsed = parse_artifact_uri(uploaded_uri)
            if parsed:
                path = await _materialize_excel_artifact(
                    tool_context,
                    parsed.filename,
                    version=parsed.version,
                    preferred_name=state.get("uploaded_excel_name", ""),
                )
                if path:
                    state["uploaded_excel_path"] = path
                    return path
        if uploaded_uri.startswith("file://"):
            local_path = _normalize_excel_path(uploaded_uri)
            if os.path.exists(local_path):
                state["uploaded_excel_path"] = local_path
                return local_path

    preferred = os.path.basename(state.get("uploaded_excel_name", "") or provided_path or "")
    try:
        artifact_names = await tool_context.list_artifacts()
    except Exception:
        artifact_names = []
    candidates = [
        name for name in artifact_names
        if os.path.splitext(name)[1].lower() in _EXCEL_EXTS
    ]
    if preferred:
        for name in candidates:
            if os.path.basename(name) == preferred:
                path = await _materialize_excel_artifact(
                    tool_context, name, preferred_name=preferred
                )
                if path:
                    state["uploaded_excel_path"] = path
                    return path
    if len(candidates) == 1:
        path = await _materialize_excel_artifact(
            tool_context, candidates[0], preferred_name=candidates[0]
        )
        if path:
            state["uploaded_excel_path"] = path
            return path

    return ""


def _resolve_excel_path(path: str) -> str:
    """Resolve a missing path by searching common output locations."""
    if os.path.exists(path):
        return path

    base = os.path.basename(path)
    if not base:
        return path

    output_dir = os.path.expanduser(OUTPUT_DIR)
    if not os.path.isabs(output_dir):
        output_dir = os.path.join(APP_ROOT, output_dir)
    output_dir = os.path.abspath(output_dir)

    root_outputs = os.path.join(APP_ROOT, "outputs")
    agent_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(output_dir, base),
        os.path.join(os.getcwd(), base),
        os.path.join(APP_ROOT, base),
        os.path.join(root_outputs, base),
        os.path.join(agent_dir, base),
        os.path.join(agent_dir, "outputs", base),
    ]
    for cand in candidates:
        if os.path.exists(cand):
            return cand

    # Fallback: search by glob pattern
    import glob as glob_module
    search_dirs = [output_dir, os.getcwd(), APP_ROOT, root_outputs, "/tmp"]
    base_name = os.path.splitext(base)[0]
    for search_dir in search_dirs:
        if not os.path.isdir(search_dir):
            continue
        for pattern in [f"{base}*", f"*{base_name}*.xlsx"]:
            matches = glob_module.glob(os.path.join(search_dir, pattern))
            if matches:
                # Return first match
                return matches[0]

    return path


def _normalize_template_id(value: str) -> str:
    """Extract the Slides template file ID from a URL or return raw ID."""
    v = (value or "").strip()
    if not v:
        return ""

    v = v.split("#", 1)[0].split("?", 1)[0]

    # Common URL formats:
    # https://docs.google.com/presentation/d/<ID>/edit
    # https://drive.google.com/file/d/<ID>/view
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", v)
    if m:
        return m.group(1)

    # id=<ID>
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", v)
    if m:
        return m.group(1)

    if "/" in v:
        return v.split("/", 1)[0]

    return v


def _convert_xls_to_xlsx_with_soffice(xls_path: str) -> tuple[Optional[str], Optional[str], str]:
    """Convert legacy .xls to .xlsx (preserving formatting) using LibreOffice."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None, None, (
            "Legacy .xls detected, but LibreOffice (soffice) is not installed/available. "
            "Please save the file as .xlsx, or install LibreOffice so this tool can convert it."
        )

    out_dir = tempfile.mkdtemp(prefix="xls_to_xlsx_")
    try:
        proc = subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                out_dir,
                xls_path,
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if proc.returncode != 0:
            return None, out_dir, f"LibreOffice conversion failed: {proc.stderr.strip() or proc.stdout.strip()}"

        base = os.path.splitext(os.path.basename(xls_path))[0]
        expected = os.path.join(out_dir, f"{base}.xlsx")
        if os.path.exists(expected):
            return expected, out_dir, ""

        # Fallback: pick the first .xlsx in the output directory
        for fname in os.listdir(out_dir):
            if fname.lower().endswith(".xlsx"):
                return os.path.join(out_dir, fname), out_dir, ""

        return None, out_dir, "LibreOffice conversion did not produce an .xlsx output."
    except Exception as exc:
        return None, out_dir, f"LibreOffice conversion error: {exc}"


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — TOOL 1: EXCEL DATA EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════

async def extract_excel_data(excel_file_path: str, *, tool_context: ToolContext) -> dict:
    """
    Parse the Excel weekly report file and extract:
      - Project metadata (project name, client, vendor, SPOC, dates)
      - Executive summary text
      - Task list from Sheet1
      - Dynamic range of the Timeline sheet

    Args:
        excel_file_path: Path/URI to an Excel file (.xlsx/.xlsm/.xltx/.xltm), 
                        legacy .xls, or Google Sheets URL.

    Returns:
        success (bool), project_name (str), placeholder_count (int),
        timeline_range (str), error (str)
    """
    import re
    
    project_name_from_msg = None
    match = re.search(r"\(Project Name:\s*([^)]+)\)", excel_file_path)
    if match:
        project_name_from_msg = match.group(1).strip()
        excel_file_path = excel_file_path.replace(match.group(0), "").strip()
        print(f"[DEBUG] Extracted project name from message: {project_name_from_msg}")
    
    # Check if it's a Google Sheets URL - just store URL, don't download
    path = ""
    if "docs.google.com/spreadsheets" in (excel_file_path or ""):
        print(f"[DEBUG] Detected Google Sheets URL - will read directly via API")
        
        # Store the Google Sheet URL in state for later use
        tool_context.state["google_sheet_url"] = excel_file_path
        tool_context.state["excel_file_path"] = ""  # No local file
    else:
        path = _resolve_excel_path(_normalize_excel_path(excel_file_path))
    
    print(f"[DEBUG extract_excel_data] original_path={excel_file_path}, resolved_path={path}")
    
    # Only check path existence for local files
    if path and not os.path.exists(path):
        fallback = await _resolve_excel_from_state(tool_context, excel_file_path)
        print(f"[DEBUG extract_excel_data] fallback_path={fallback}")
        if fallback:
            path = fallback

    # For Google Sheets URLs, skip local file processing
    if "docs.google.com/spreadsheets" in (excel_file_path or ""):
        print("[DEBUG] Skipping local file processing for Google Sheets")
        placeholders = {
            "{{Project Name}}": "Weekly Report",
            "{{From Date}}": "",
            "{{To Date}}": "",
        }
        tool_context.state["placeholder_map"] = placeholders
        tool_context.state["project_name"] = "Weekly Report"
        tool_context.state["timeline_range"] = "A1:Z50"
        tool_context.state["timeline_error"] = ""
        tool_context.state["tasks_data"] = []
        return {
            "success": True,
            "project_name": "Weekly Report",
            "placeholder_count": len(placeholders),
            "timeline_range": "A1:Z50",
            "timeline_error": "",
            "error": "",
        }

    if "fakepath" in (excel_file_path or "").lower() and not os.path.exists(path):
        return {
            "success": False,
            "project_name": "",
            "placeholder_count": 0,
            "timeline_range": "",
            "error": (
                f"File not found: {path}. "
                "It looks like a browser 'C:\\fakepath\\...' value was provided. "
                "Please provide the server-side file path (e.g. the uploaded temp_path)."
            ),
        }

    if not os.path.exists(path):
        return {
            "success": False,
            "project_name": "",
            "placeholder_count": 0,
            "timeline_range": "",
            "error": f"File not found: {path}",
        }

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        converted_path, converted_dir, err = _convert_xls_to_xlsx_with_soffice(path)
        if not converted_path:
            return {
                "success": False,
                "project_name": "",
                "placeholder_count": 0,
                "timeline_range": "",
                "error": err,
            }
        tool_context.state["converted_excel_path"] = converted_path
        tool_context.state["converted_excel_dir"] = converted_dir
        path = converted_path
        ext = ".xlsx"

    supported = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if ext not in supported:
        return {
            "success": False,
            "project_name": "",
            "placeholder_count": 0,
            "timeline_range": "",
            "error": f"Unsupported Excel file type: {ext}. Supported: {', '.join(sorted(supported))} (and .xls with LibreOffice).",
        }

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        return {
            "success": False,
            "project_name": "",
            "placeholder_count": 0,
            "timeline_range": "",
            "error": f"Cannot open workbook: {exc}",
        }

    # ── Extract metadata from workbook ────────────────────────────────────
    sheet1 = wb.active
    meta   = _extract_metadata(wb, sheet1)

    # ── Extract task rows from workbook ───────────────────────────────────
    tasks  = _extract_tasks(wb, sheet1)
    print(f"[DEBUG] Tasks extracted: {len(tasks)}")
    if tasks:
        print(f"[DEBUG] First task sample: {tasks[0]}")

    # ── Dynamic Timeline range detection ──────────────────────────────────
    timeline_range = ""
    timeline_error = ""
    tl_ws = None
    print(f"[DEBUG] Available sheets: {wb.sheetnames}")
    print(f"[DEBUG] Looking for: '{TIMELINE_SHEET_NAME}'")
    
    # Exact match first
    if TIMELINE_SHEET_NAME in wb.sheetnames:
        tl_ws = wb[TIMELINE_SHEET_NAME]
        print(f"[DEBUG] Exact match found: {TIMELINE_SHEET_NAME}")
    else:
        # Try with stripped whitespace
        for name in wb.sheetnames:
            if name.strip() == TIMELINE_SHEET_NAME.strip():
                tl_ws = wb[name]
                print(f"[DEBUG] Stripped match found: {name}")
                break
        else:
            # Fallback: try to locate a timeline/gantt sheet by name
            for name in wb.sheetnames:
                lname = name.lower()
                if "timeline" in lname or "gantt" in lname or "plan" in lname:
                    tl_ws = wb[name]
                    print(f"[DEBUG] Fuzzy match found: {name}")
                    break

    if tl_ws is not None:
        timeline_range = _get_dynamic_range(tl_ws)
        print(f"[DEBUG] Timeline sheet found, range: {timeline_range}")
    else:
        available      = ", ".join(wb.sheetnames)
        timeline_error = (
            f"Sheet '{TIMELINE_SHEET_NAME}' not found. "
            f"Available sheets: {available}"
        )
        print(f"[DEBUG] Timeline sheet NOT found. {timeline_error}")

    print(f"[DEBUG] Meta before: {meta}")
    print(f"[DEBUG] Tasks count: {len(tasks)}")

    # ── Persist to session state ──────────────────────────────────────────
    # Use Excel filename as project name (without extension)
    file_name = os.path.basename(path)
    project_name_from_file = os.path.splitext(file_name)[0] if file_name else "Project"
    if project_name_from_file.startswith("upload_"):
        project_name_from_file = "Project"
    if project_name_from_msg:
        project_name_from_file = project_name_from_msg
    
    # Update meta with filename so placeholder map gets correct value
    meta["project_name"] = project_name_from_file
    placeholders = _build_placeholder_map(meta, tasks)
    
    tool_context.state["excel_file_path"]  = path
    tool_context.state["placeholder_map"]  = placeholders
    tool_context.state["project_name"]     = project_name_from_file
    tool_context.state["timeline_range"]     = timeline_range
    tool_context.state["timeline_error"]   = timeline_error
    tool_context.state["tasks_data"]       = tasks
    tool_context.state["debug_meta"]       = meta
    tool_context.state["debug_tasks"]      = tasks[:5] if tasks else []

    return {
        "success":           True,
        "project_name":      meta.get("project_name", ""),
        "placeholder_count": len(placeholders),
        "timeline_range":    timeline_range,
        "timeline_error":    timeline_error,
        "debug_tasks_sample": str(tasks[:3]) if tasks else "[]",
        "debug_meta":        meta,
        "debug_placeholder_values": {k: str(v)[:50] for k, v in placeholders.items()},
        "error":             "",
    }


def _get_dynamic_range(ws) -> str:
    """
    Scan every cell in the sheet and find the true last row and column
    that contains data OR a coloured cell background (Gantt bar cells
    often have no text but do have a fill colour).
    Never assume a fixed boundary — always recalculate.
    """
    max_row = 0
    max_col = 0
    
    # Check sheet dimensions first - use a larger default since Gantt charts often span many rows
    sheet_max_row = max(ws.max_row or 1, 50)  # Minimum 50 rows for Gantt
    sheet_max_col = max(ws.max_column or 1, 10)  # Minimum 10 columns
    
    print(f"[DEBUG _get_dynamic_range] Sheet dimensions: {sheet_max_col}x{sheet_max_row}")
    
    # Check for tables in the workbook
    if hasattr(ws, 'tables'):
        print(f"[DEBUG _get_dynamic_range] Tables found: {list(ws.tables.keys())}")
    
    # Debug: sample some cells
    sample_cells = []
    for row in list(ws.iter_rows(max_row=5)):
        for cell in row[:8]:  # First 8 columns
            has_value = cell.value is not None and str(cell.value).strip() != ""
            has_color = _cell_bg_rgb(cell) is not None
            if has_value or has_color:
                sample_cells.append(f"{cell.coordinate}: val={cell.value}, color={has_color}")
    
    print(f"[DEBUG _get_dynamic_range] Sample cells with content: {sample_cells[:20]}")

    for row in ws.iter_rows():
        for cell in row:
            has_value = cell.value is not None and str(cell.value).strip() != ""
            has_color = _cell_bg_rgb(cell) is not None
            has_border = False
            try:
                b = cell.border
                has_border = any([
                    b.left   and b.left.style,
                    b.right  and b.right.style,
                    b.top    and b.top.style,
                    b.bottom and b.bottom.style,
                ])
            except Exception:
                pass

            if has_value or has_color or has_border:
                if cell.row    > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column

    # Fallback to sheet dimensions if detection is too small
    # For Gantt charts, default to at least 30 rows and 10 columns
    if max_row < 5:
        max_row = max(sheet_max_row, 30)
    if max_col < 2:
        max_col = max(sheet_max_col, 10)
        
    print(f"[DEBUG _get_dynamic_range] Final: {max_col}x{max_row}")
    last_col_letter = get_column_letter(max_col)
    return f"A1:{last_col_letter}{max_row}"


def _extract_metadata(wb, sheet1) -> dict:
    """
    Pull project metadata from the workbook.
    Looks for key-value pairs in the first few rows across sheets.
    """
    meta = {
        "project_name":  "",
        "client_name":   "",
        "vendor_name":   "",
        "spoc_name":     "",
        "spoc_email":    "",
        "start_date":    "",
        "end_date":      "",
        "total_weeks":   "",
        "report_date":   datetime.today().strftime("%d %B %Y"),
    }

    def _apply_kv(key_raw: str, val: str) -> None:
        if "project" in key_raw and "name" in key_raw and not meta["project_name"]:
            meta["project_name"] = _na(val)
        elif "client" in key_raw and not meta["client_name"]:
            meta["client_name"]  = _na(val)
        elif "vendor" in key_raw and not meta["vendor_name"]:
            meta["vendor_name"]  = _na(val)
        elif "spoc" in key_raw and "email" not in key_raw and not meta["spoc_name"]:
            meta["spoc_name"]    = _na(val)
        elif "spoc" in key_raw and "email" in key_raw and not meta["spoc_email"]:
            meta["spoc_email"]   = _na(val)
        elif "start" in key_raw and "date" in key_raw and not meta["start_date"]:
            meta["start_date"]   = _na(val)
        elif "end" in key_raw and "date" in key_raw and not meta["end_date"]:
            meta["end_date"]     = _na(val)
        elif "week" in key_raw and not meta["total_weeks"]:
            meta["total_weeks"]  = _na(val)

    # Also try to get from first few rows of Task Tracker (column A/B as key/value)
    # Skip rows that look like column headers
    header_keywords = {"phase", "module", "tasks", "start date", "end date", "status", "remarks", "duration", "assigned", "start", "end"}
    if "Task Tracker" in wb.sheetnames:
        ts = wb["Task Tracker"]
        for row_i, row in enumerate(ts.iter_rows(max_row=10, values_only=True), start=1):
            if not row or len(row) < 2:
                continue
            key_cell = str(row[0]).strip().lower() if row[0] else ""
            val_cell = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not key_cell or key_cell in header_keywords:
                continue
            print(f"[DEBUG _extract_metadata] Found key={row[0]}, val={val_cell}")
            _apply_kv(key_cell, val_cell)

    # Scan all cells for keyword patterns across sheets
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=min(40, ws.max_row), values_only=True):
            for i, cell_val in enumerate(row):
                if cell_val is None:
                    continue
                raw = str(cell_val).strip()
                if not raw:
                    continue
                key_raw = raw.lower()

                val = ""
                if ":" in raw:
                    key_part, val_part = raw.split(":", 1)
                    key_raw = key_part.strip().lower()
                    val = val_part.strip()
                else:
                    # Next cell in row = likely value
                    val = str(row[i + 1]).strip() if i + 1 < len(row) and row[i + 1] else ""

                _apply_kv(key_raw, val)

    # Derive project name from workbook title if still empty
    if not meta["project_name"] or meta["project_name"] == "N/A":
        if wb.properties.title:
            meta["project_name"] = wb.properties.title
        else:
            meta["project_name"] = "Weekly Report"

    return meta


_TASK_HEADER_MAP = {
    "phase":       ["phase", "stage", "track"],
    "module":      ["module", "workstream", "area", "work stream"],
    "task_detail": ["task", "task detail", "task details", "activity", "description", "scope"],
    "assigned_to": ["assigned", "assignee", "owner", "responsible", "resource"],
    "start_date":  ["start", "start date", "from"],
    "end_date":    ["end", "end date", "to", "due"],
    "status":      ["status", "state", "progress"],
    "remark":      ["remark", "remarks", "comment", "notes"],
}


def _canonical_header(header: str) -> str:
    h = header.strip().lower()
    for canon, keys in _TASK_HEADER_MAP.items():
        if any(k in h for k in keys):
            return canon
    return h


def _score_header_row(headers: list[str]) -> int:
    score = 0
    for h in headers:
        if not h:
            continue
        for keys in _TASK_HEADER_MAP.values():
            if any(k in h for k in keys):
                score += 1
                break
    return score


def _extract_tasks(wb, sheet1) -> list:
    """Extract task rows from the workbook (Task Tracker format)."""
    tasks = []

    best = {"score": 0, "sheet": sheet1, "row_i": 1, "headers": []}
    for ws in wb.worksheets:
        for row_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_i > 40:
                break
            headers = [str(c).strip().lower() if c else "" for c in row]
            score = _score_header_row(headers)
            if score > best["score"]:
                best = {"score": score, "sheet": ws, "row_i": row_i, "headers": headers}

    ws = best["sheet"]
    header_row = best["row_i"]
    headers = best["headers"] or []
    if not any(headers):
        # Fallback to first row of active sheet
        ws = sheet1
        header_row = 1
        headers = [str(c).strip().lower() if c else "" for c in next(ws.iter_rows(values_only=True), [])]

    headers = [_canonical_header(h) for h in headers]

    for row_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_i <= header_row:
            continue
        if all(c is None for c in row):
            continue

        task = {}
        for col_i, val in enumerate(row):
            if col_i < len(headers) and headers[col_i]:
                task[headers[col_i]] = _na(val)
        if any(v and v != "N/A" for v in task.values()):
            tasks.append(task)

    return tasks


def _build_placeholder_map(meta: dict, tasks: list) -> dict:
    """
    Build the {{PLACEHOLDER}} → replacement text mapping.
    All template tags should be listed here.
    """
    
    # Task status summary
    def _status_value(t: dict) -> str:
        for key in ("status", "state", "progress"):
            if key in t:
                return str(t.get(key, ""))
        return ""

    def _parse_date(val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val
        raw = str(val).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y"):
            try:
                return datetime.strptime(raw.split()[0], fmt.split()[0] if " " in fmt else fmt)
            except:
                continue
        return None

    done    = sum(1 for t in tasks if "complet" in _status_value(t).lower())
    inprog  = sum(1 for t in tasks if "progress" in _status_value(t).lower())
    todo    = sum(1 for t in tasks if "todo" in _status_value(t).lower()
                                  or "to do" in _status_value(t).lower())
    delayed = sum(1 for t in tasks if "delay" in _status_value(t).lower())
    total   = len(tasks)
    
    # Derive dates from tasks if not in meta
    start_dates = []
    end_dates = []
    for t in tasks:
        sd = _parse_date(t.get("start_date"))
        ed = _parse_date(t.get("end_date"))
        if sd:
            start_dates.append(sd)
        if ed:
            end_dates.append(ed)
    
    earliest_start = min(start_dates) if start_dates else None
    latest_end = max(end_dates) if end_dates else None
    
    start_date_str = meta.get("start_date", "")
    end_date_str = meta.get("end_date", "")
    
    # Override if dates come from tasks and meta has invalid values
    if earliest_start and (not start_date_str or start_date_str in ("End Date", "Duration", "N/A")):
        start_date_str = earliest_start.strftime("%d %b %Y")
    if latest_end and (not end_date_str or end_date_str in ("End Date", "Duration", "N/A")):
        end_date_str = latest_end.strftime("%d %b %Y")
    
    # Calculate weeks
    total_weeks = meta.get("total_weeks", "")
    if not total_weeks and earliest_start and latest_end:
        weeks = (latest_end - earliest_start).days // 7
        if weeks > 0:
            total_weeks = str(weeks)
    
    print(f"[DEBUG _build_placeholder_map] meta={meta}, total_tasks={total}, done={done}, inprog={inprog}, todo={todo}, start_date={start_date_str}, end_date={end_date_str}")

    return {
        # Template placeholders (exact format)
        "{{Project Name}}":   _na(meta.get("project_name")),
        "{{From Date}}":      _na(start_date_str),
        "{{To Date}}":        _na(end_date_str),
    }


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 — TOOL 2: TIMELINE SCREENSHOT RENDERER
# ══════════════════════════════════════════════════════════════════════════════

def capture_timeline_screenshot(*, tool_context: ToolContext) -> dict:
    """
    Render the Timeline sheet range to a high-resolution PNG using
    openpyxl + Pillow, then upload it to Google Drive.

    Reads session state:
        excel_file_path  (str)  — path to the .xlsx file
        timeline_range   (str)  — e.g. "A1:M32"

    Writes session state:
        drive_image_id   (str)  — Google Drive file ID of the uploaded PNG
        drive_image_url  (str)  — public/direct URL
        image_width_px   (int)
        image_height_px  (int)

    Returns:
        success (bool), drive_image_id (str), width_px (int),
        height_px (int), error (str)
    """
    state          = tool_context.state
    excel_path     = state.get("excel_file_path", "")
    timeline_range = state.get("timeline_range", "")
    google_sheet_url = state.get("google_sheet_url", "")
    
    print(f"[DEBUG screenshot] excel_path={excel_path}, google_sheet_url={google_sheet_url}")

    # For Google Sheets, image will be generated in build_slides
    if google_sheet_url and "docs.google.com/spreadsheets" in google_sheet_url:
        print("[DEBUG] Google Sheets URL - skipping screenshot, image will be created in build_slides")
        state["drive_image_id"] = ""
        state["image_width_px"] = 0
        state["image_height_px"] = 0
        return {"success": True, "drive_image_id": "", "width_px": 0, "height_px": 0, "error": ""}

    if not excel_path or not os.path.exists(excel_path):
        return {"success": False, "drive_image_id": "", "width_px": 0,
                "height_px": 0, "error": f"Excel file not found: {excel_path}"}

    if not timeline_range:
        return {"success": False, "drive_image_id": "", "width_px": 0,
                "height_px": 0,
                "error": "No timeline_range in state. Run extract_excel_data first."}

    # ── Upload Excel to Google Drive as Sheets ─────────────────────────────
    try:
        drive, slides = _build_services()
        sheets_service = build("sheets", "v4", credentials=_get_credentials())
    except Exception as exc:
        return {"success": False, "drive_image_id": "", "width_px": 0,
                "height_px": 0, "error": f"Failed to build services: {exc}"}

    try:
        # Upload Excel file to Drive
        file_meta = {
            "name":     f"timeline_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        if PARENT_FOLDER_ID:
            file_meta["parents"] = [PARENT_FOLDER_ID]

        media = MediaFileUpload(excel_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resumable=False)
        uploaded = drive.files().create(
            body=file_meta, media_body=media,
            fields="id,name",
            supportsAllDrives=True,
        ).execute()
        
        spreadsheet_id = uploaded.get("id")
        print(f"[DEBUG] Uploaded spreadsheet ID: {spreadsheet_id}")
        
        # Make it readable
        drive.permissions().create(
            fileId=spreadsheet_id,
            body={"type": "anyone", "role": "reader"},
            supportsAllDrives=True,
        ).execute()
        
        # Get data from Timeline sheet using Sheets API
        try:
            # Try to get Timeline sheet data
            result = sheets_service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=f"'{TIMELINE_SHEET_NAME}'!A1:Z50",
            ).execute()
            timeline_data = result.get("values", [])
            print(f"[DEBUG] Timeline sheet data rows: {len(timeline_data)}")
        except Exception as e:
            print(f"[DEBUG] Could not read Timeline sheet: {e}")
            timeline_data = []
        
        # If no timeline data, use task data from state
        if not timeline_data:
            tasks_data = state.get("tasks_data", [])
            if tasks_data:
                # Build a simple table from tasks
                timeline_data = [["Task", "Start Date", "End Date", "Status"]]
                for task in tasks_data[:15]:  # Limit to 15 rows
                    timeline_data.append([
                        task.get("task_detail", "")[:30],
                        task.get("start_date", ""),
                        task.get("end_date", ""),
                        task.get("status", "")
                    ])
                print(f"[DEBUG] Using task data for timeline: {len(timeline_data)-1} tasks")
        
        # Generate image from the data
        if timeline_data:
            img = _render_table_to_png(timeline_data, dpi=150)
        else:
            return {"success": False, "drive_image_id": "", "width_px": 0,
                    "height_px": 0, "error": "No timeline data found"}
        
    except Exception as exc:
        import traceback
        return {"success": False, "drive_image_id": "", "width_px": 0,
                "height_px": 0,
                "error": f"Upload/Sheets API failed: {exc}\n{traceback.format_exc()}"}

    w_px, h_px = img.size

    # ── Save to temp file ─────────────────────────────────────────────────
    tmp_fd, tmp_path = tempfile.mkstemp(suffix="_timeline.png")
    os.close(tmp_fd)
    img.save(tmp_path, "PNG", dpi=(150, 150))

    # ── Upload to Google Drive ─────────────────────────────────────────────
    try:
        drive, _ = _build_services()
        file_meta = {
            "name":     f"timeline_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
            "mimeType": "image/png",
        }
        if PARENT_FOLDER_ID:
            file_meta["parents"] = [PARENT_FOLDER_ID]

        media = MediaFileUpload(tmp_path, mimetype="image/png", resumable=False)
        uploaded = drive.files().create(
            body=file_meta, media_body=media,
            fields="id,webContentLink",
            supportsAllDrives=True,
        ).execute()

        drive_image_id  = uploaded.get("id", "")
        drive_image_url = uploaded.get(
            "webContentLink",
            f"https://drive.google.com/uc?id={drive_image_id}"
        )

        # Make image readable by the Slides API (anyone with link)
        drive.permissions().create(
            fileId=drive_image_id,
            body={"type": "anyone", "role": "reader"},
            supportsAllDrives=True,
        ).execute()

    except HttpError as exc:
        os.remove(tmp_path)
        return {"success": False, "drive_image_id": "", "width_px": 0,
                "height_px": 0, "error": f"Drive upload failed: {exc}"}
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    state["drive_image_id"]  = drive_image_id
    state["drive_image_url"] = drive_image_url
    state["image_width_px"]  = w_px
    state["image_height_px"] = h_px
    state["spreadsheet_id"] = spreadsheet_id  # For cleanup

    return {"success": True, "drive_image_id": drive_image_id,
            "width_px": w_px, "height_px": h_px, "error": ""}


def _render_table_to_png(table_data: list, dpi: int = 150) -> Image.Image:
    """
    Render a 2D list as a table image.
    """
    if not table_data:
        return Image.new("RGB", (100, 100), (255, 255, 255))
    
    SCALE = dpi / 96.0
    PX_PER_PT = dpi / 72.0
    
    # Cell dimensions
    col_width = int(120 * SCALE)
    row_height = int(25 * PX_PER_PT)
    header_height = int(30 * PX_PER_PT)
    
    # Calculate dimensions
    num_cols = max(len(row) for row in table_data)
    num_rows = len(table_data)
    
    total_w = num_cols * col_width
    total_h = header_height + (num_rows - 1) * row_height
    
    img = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    
    # Try to load fonts
    try:
        font_header = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", max(10, int(12 * SCALE)))
        font_normal = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", max(9, int(10 * SCALE)))
    except Exception:
        font_header = ImageFont.load_default()
        font_normal = font_header
    
    # Draw header row
    header_fill = (66, 133, 244)  # Blue
    draw.rectangle([0, 0, total_w, header_height], fill=header_fill)
    
    # Draw header text
    header_row = table_data[0]
    for col_idx, cell in enumerate(header_row):
        x = col_idx * col_width + 5
        y = 5
        draw.text((x, y), str(cell)[:20], fill=(255, 255, 255), font=font_header)
    
    # Draw data rows
    for row_idx, row in enumerate(table_data[1:], start=1):
        y0 = header_height + (row_idx - 1) * row_height
        
        # Alternate row colors
        if row_idx % 2 == 0:
            row_fill = (245, 245, 245)
        else:
            row_fill = (255, 255, 255)
        draw.rectangle([0, y0, total_w, y0 + row_height], fill=row_fill)
        
        # Draw cell text
        for col_idx, cell in enumerate(row):
            if col_idx >= num_cols:
                break
            x = col_idx * col_width + 5
            y = y0 + 5
            draw.text((x, y), str(cell)[:20], fill=(0, 0, 0), font=font_normal)
        
        # Row border
        draw.line([(0, y0), (total_w, y0)], fill=(200, 200, 200), width=1)
    
    # Column borders
    for col_idx in range(num_cols):
        x = col_idx * col_width
        draw.line([(x, 0), (x, total_h)], fill=(200, 200, 200), width=1)
    
    # Outer border
    draw.rectangle([0, 0, total_w - 1, total_h - 1], outline=(100, 100, 100), width=2)
    
    return img


def _render_gantt_chart(tasks: list, dpi: int = 150) -> Image.Image:
    import traceback
    from datetime import timedelta
    from matplotlib.patches import FancyBboxPatch
    from matplotlib.dates import date2num
    print(f"[DEBUG GANTT] Starting with {len(tasks)} tasks")
    try:
        if not tasks:
            print("[DEBUG GANTT] No tasks - returning empty image")
            return Image.new("RGB", (800, 400), (255, 255, 255))
        
        phases = []
        module_names = []
        task_names = []
        start_dates = []
        end_dates = []
        statuses = []
        
        def parse_date(val):
            if not val:
                return None
            if isinstance(val, datetime):
                return val
            raw = str(val).strip()
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y"):
                try:
                    return datetime.strptime(raw.split()[0], fmt)
                except:
                    continue
            return None
        
        for task in tasks[:15]:
            task_name = str(task.get("task_detail", "")).strip()[:20] or str(task.get("module", "")).strip()[:20]
            module = str(task.get("module", "")).strip()
            phase = str(task.get("phase", "")).strip()[:15]
            if not task_name:
                continue
                
            start = parse_date(task.get("start_date", ""))
            end = parse_date(task.get("end_date", ""))
            status = str(task.get("status", "")).lower()
            
            if start and end:
                phases.append(phase)
                module_names.append(module)
                task_names.append(task_name)
                start_dates.append(start)
                end_dates.append(end)
                statuses.append(status)
        
        if not task_names:
            print(f"[DEBUG GANTT] No valid tasks with dates - task_names empty. Parsed: {len(phases)} phases")
            return Image.new("RGB", (800, 400), (255, 255, 255))
        
        fig, ax = plt.subplots(figsize=(14, max(6, len(task_names) * 0.6)))
        fig.set_facecolor('#f8f9fa')
        ax.set_facecolor('#f8f9fa')
        
        if start_dates and end_dates:
            x_min = date2num(min(start_dates))
            x_max = date2num(max(end_dates))
            ax.set_xlim(x_min - 1, x_max + 1)
        
        color_map = {
            "completed": "#2ecc71",
            "progress": "#f39c12",
            "in progress": "#f39c12",
            "todo": "#3498db",
            "to do": "#3498db",
            "delayed": "#e74c3c",
        }
        
        for i, (phase, module, start, end, status) in enumerate(zip(phases, module_names, start_dates, end_dates, statuses)):
            color = color_map.get(status, "#3498db")
            start_num = date2num(start)
            width = (end - start).days
            bar = FancyBboxPatch((start_num, i - 0.325), width, 0.65, boxstyle="round,pad=0.02,rounding_size=0.15", facecolor=color, alpha=0.9, edgecolor='none')
            ax.add_patch(bar)
            ax.text(start_num + 0.5, i, module, va='center', ha='left', fontsize=9, color='black')
        
        ax.set_ylim(-0.5, len(phases) - 0.5)
        ax.set_yticks(range(len(phases)))
        ax.set_yticklabels(phases, fontsize=10, fontweight='medium', color='#2c3e50')
        ax.set_xlabel("")
        ax.xaxis.set_label_position('top')
        ax.set_title("Project Timeline", fontsize=17, fontweight='bold', color='#2c3e50', pad=15)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d %b'))
        ax.xaxis.set_major_locator(mdates.DayLocator())
        ax.tick_params(axis='x', top=True, labeltop=True, bottom=False, labelbottom=False, rotation=45, labelsize=8, colors='#7f8c8d')
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, axis='x', linestyle='-', alpha=0.2, color='#bdc3c7')
        ax.set_axisbelow(True)
        plt.tight_layout(pad=1.5)
        
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
        os.close(tmp_fd)
        plt.savefig(tmp_path, dpi=dpi, bbox_inches='tight')
        plt.close()
        
        img = Image.open(tmp_path)
        os.remove(tmp_path)
        print(f"[DEBUG GANTT] Chart created successfully: {img.size}")
        return img
    except Exception as e:
        print(f"[DEBUG GANTT] Error creating Gantt chart: {e}")
        traceback.print_exc()
        return Image.new("RGB", (800, 400), (255, 255, 255))


def _render_sheet_range_to_png(
    excel_path: str,
    sheet_name: str,
    cell_range: str,
    dpi: int = 150,
) -> Image.Image:
    """
    Render an openpyxl worksheet range to a PIL Image.

    Note: This renders cell content only, not embedded images.
    If the Timeline sheet has pasted images, they won't be captured.
    
    Strategy:
      • Map each cell's width/height from openpyxl column/row dimensions
      • Fill cell background with its actual fill colour
      • Draw cell borders
      • Render cell text with basic font sizing
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # Parse range  e.g. "A1:M32"
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    SCALE     = dpi / 96.0          # 96 dpi is openpyxl's reference
    PX_PER_PT = dpi / 72.0          # points → pixels

    # ── Compute pixel dimensions per column and row ───────────────────────
    DEFAULT_COL_W_PX = int(8.43 * 7 * SCALE)   # openpyxl default: 8.43 chars ≈ 64px
    DEFAULT_ROW_H_PX = int(15  * PX_PER_PT)     # openpyxl default: 15pt

    col_widths = {}
    for col_i in range(min_col, max_col + 1):
        letter = get_column_letter(col_i)
        dim    = ws.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths[col_i] = max(4, int(dim.width * 7 * SCALE))
        else:
            col_widths[col_i] = DEFAULT_COL_W_PX

    row_heights = {}
    for row_i in range(min_row, max_row + 1):
        dim = ws.row_dimensions.get(row_i)
        if dim and dim.height:
            row_heights[row_i] = max(4, int(dim.height * PX_PER_PT))
        else:
            row_heights[row_i] = DEFAULT_ROW_H_PX

    total_w = sum(col_widths[c]  for c in range(min_col, max_col + 1))
    total_h = sum(row_heights[r] for r in range(min_row, max_row + 1))

    img  = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # Try to load a small font; fall back to default if unavailable
    try:
        font_normal = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                                         max(8, int(10 * SCALE)))
        font_bold   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                                         max(8, int(10 * SCALE)))
    except Exception:
        font_normal = ImageFont.load_default()
        font_bold   = font_normal

    # ── Pre-compute column X offsets ─────────────────────────────────────
    col_x = {}
    x = 0
    for c in range(min_col, max_col + 1):
        col_x[c] = x
        x += col_widths[c]

    # ── Pre-compute row Y offsets ─────────────────────────────────────────
    row_y = {}
    y = 0
    for r in range(min_row, max_row + 1):
        row_y[r] = y
        y += row_heights[r]

    # ── Draw cells ────────────────────────────────────────────────────────
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            x0   = col_x[c]
            y0   = row_y[r]
            x1   = x0 + col_widths[c]
            y1   = y0 + row_heights[r]

            # Background fill
            bg = _cell_bg_rgb(cell)
            if bg:
                draw.rectangle([x0, y0, x1 - 1, y1 - 1], fill=bg)
            else:
                draw.rectangle([x0, y0, x1 - 1, y1 - 1], fill=(255, 255, 255))

            # Cell text
            text = str(cell.value).strip() if cell.value is not None else ""
            if text:
                # Font colour
                try:
                    fc = cell.font.color
                    if fc and fc.type == "rgb" and fc.rgb:
                        txt_color = _hex_to_rgb(fc.rgb)
                    else:
                        txt_color = (0, 0, 0)
                except Exception:
                    txt_color = (0, 0, 0)

                is_bold = False
                try:
                    is_bold = bool(cell.font.bold)
                except Exception:
                    pass

                fnt     = font_bold if is_bold else font_normal
                padding = max(2, int(3 * SCALE))

                # Truncate text to fit cell width
                max_text_w = col_widths[c] - 2 * padding
                while text:
                    bbox = draw.textbbox((0, 0), text, font=fnt)
                    if (bbox[2] - bbox[0]) <= max_text_w:
                        break
                    text = text[:-1]
                    if len(text) > 3:
                        text = text[:-3] + "…"
                    else:
                        text = text[:-1]

                if text:
                    draw.text(
                        (x0 + padding, y0 + padding),
                        text,
                        fill=txt_color,
                        font=fnt,
                    )

            # Border lines (thin light grey)
            BORDER_COLOR = (200, 200, 200)
            draw.line([(x0, y0), (x1, y0)],         fill=BORDER_COLOR, width=1)
            draw.line([(x0, y0), (x0, y1)],         fill=BORDER_COLOR, width=1)
            draw.line([(x1 - 1, y0), (x1 - 1, y1)], fill=BORDER_COLOR, width=1)
            draw.line([(x0, y1 - 1), (x1, y1 - 1)], fill=BORDER_COLOR, width=1)

    return img


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 2 — TOOL 3: SLIDES BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_slides_report(*, tool_context: ToolContext) -> dict:
    """
    Duplicate the template Google Slides presentation, inject all
    {{PLACEHOLDER}} text values, and replace {{TIMELINE_SCREENSHOT}}
    with the uploaded PNG (smart-scaled to fit the slide).

    Reads session state:
        placeholder_map  (dict)  — {{TAG}} → replacement text
        project_name     (str)
        drive_image_id   (str)   — Drive file ID of the PNG
        image_width_px   (int)
        image_height_px  (int)

    Writes session state:
        slides_id  (str)
        slides_url (str)

    Returns:
        success (bool), slides_url (str), slides_id (str), error (str)
    """
    state           = tool_context.state
    placeholder_map = _safe_json(state.get("placeholder_map")) or {}
    project_name    = state.get("project_name", "Report")
    drive_image_id  = state.get("drive_image_id",  "")
    img_w           = int(state.get("image_width_px",  960))
    img_h           = int(state.get("image_height_px", 540))

    print(f"[DEBUG build_slides] placeholder_map: {placeholder_map}")
    print(f"[DEBUG build_slides] project_name: {project_name}")

    template_id = _normalize_template_id(TEMPLATE_ID)
    print(f"[DEBUG build_slides] template_id: {template_id}")
    if not template_id:
        return {"success": False, "slides_url": "", "slides_id": "",
                "error": "GOOGLE_SLIDES_TEMPLATE_ID not set in .env"}

    try:
        drive, slides = _build_services()
    except Exception as exc:
        return {"success": False, "slides_url": "", "slides_id": "",
                "error": f"Google API init failed: {exc}"}

    # Debug: fetch template to see what text it contains
    try:
        template_prs = slides.presentations().get(presentationId=template_id).execute()
        template_texts = []
        for slide in template_prs.get("slides", [])[:3]:
            for elem in slide.get("pageElements", []):
                shape = elem.get("shape", {})
                text = shape.get("text", {})
                for te in text.get("textElements", []):
                    content = te.get("textRun", {}).get("content", "")
                    if content.strip():
                        template_texts.append(content.strip()[:100])
        print(f"[DEBUG] Template text samples: {template_texts[:10]}")
    except Exception as e:
        print(f"[DEBUG] Could not fetch template: {e}")

    # ── 1. Get data from Google Sheets if URL provided in state ──────────────
    google_sheet_url = state.get("google_sheet_url", "")
    sheet_img = None  # Image from Google Sheet
    sheet_img_w = 0
    sheet_img_h = 0
    
    if google_sheet_url and "docs.google.com/spreadsheets" in google_sheet_url:
        try:
            sheets_service = build("sheets", "v4", credentials=_get_credentials())
            sheet_match = re.search(r'/d/([a-zA-Z0-9_-]+)', google_sheet_url)
            if sheet_match:
                sheet_id = sheet_match.group(1)
                
                # Get spreadsheet metadata to get the title
                sheet_meta = sheets_service.spreadsheets().get(
                    spreadsheetId=sheet_id,
                    fields="properties(title)"
                ).execute()
                sheet_title = sheet_meta.get("properties", {}).get("title", "Weekly Report")
                print(f"[DEBUG] Google Sheet title: {sheet_title}")
                
                # Store project name from sheet title
                if not project_name or project_name in ("Report", "Weekly Report"):
                    project_name = sheet_title
                
                # Get data from Task Tracker sheet
                result = sheets_service.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range="Task Tracker!A1:H50",
                ).execute()
                task_data = result.get("values", [])
                print(f"[DEBUG] Task data rows: {len(task_data)}")
                
                # Extract actual dates and project name from task data
                if task_data and len(task_data) > 1:
                    # Find start and end dates from task data
                    dates = []
                    completed_tasks = []
                    in_progress_tasks = []
                    upcoming_tasks = []
                    from datetime import datetime
                    
                    header = [str(h).strip().lower() for h in task_data[0]]
                    task_idx = next((i for i, h in enumerate(header) if "task" in h), None)
                    status_idx = next((i for i, h in enumerate(header) if "status" in h), None)
                    
                    # Parse task data to extract activities
                    for row in task_data[1:]:  # Skip header
                        if len(row) >= 4:
                            task_name = ""
                            if task_idx is not None and len(row) > task_idx and row[task_idx]:
                                task_name = str(row[task_idx]).strip()
                            if not task_name and row:
                                task_name = str(row[0]).strip()
                            status = ""
                            if status_idx is not None and len(row) > status_idx and row[status_idx]:
                                status = str(row[status_idx]).lower().strip()
                            elif row and row[-1]:
                                status = str(row[-1]).lower().strip()
                            if not task_name:
                                continue
                            
                            if "completed" in status:
                                completed_tasks.append(task_name)
                            elif "progress" in status or "in progress" in status:
                                in_progress_tasks.append(task_name)
                            else:
                                upcoming_tasks.append(task_name)
                            
                            # Extract dates - parse and track min/max
                            parsed_dates = []
                            for cell in row:
                                if cell:
                                    cell_str = str(cell).strip()
                                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
                                        try:
                                            d = datetime.strptime(cell_str, fmt)
                                            parsed_dates.append(d)
                                            break
                                        except:
                                            pass
                            
                            if parsed_dates:
                                min_date = min(parsed_dates)
                                max_date = max(parsed_dates)
                                if not dates:
                                    dates = [min_date, max_date]
                                else:
                                    if min_date < dates[0]:
                                        dates[0] = min_date
                                    if max_date > dates[-1]:
                                        dates[-1] = max_date
                    
                    # Generate executive summary using LLM
                    exec_summary = _generate_executive_summary_llm(
                        completed_tasks, in_progress_tasks, upcoming_tasks
                    )
                    
                    # ── Build structured two-level bullet content with LLM-generated headings ──────────────
                    # Use LLM to create short headings from tasks (without rephrasing tasks)
                    
                    # Prepare task lists for heading generation
                    completed_raw = "\n".join(f"{i+1}. {t}" for i, t in enumerate(completed_tasks[:5])) if completed_tasks else ""
                    in_progress_raw = "\n".join(f"{i+1}. {t}" for i, t in enumerate(in_progress_tasks[:5])) if in_progress_tasks else ""
                    upcoming_raw = "\n".join(f"{i+1}. {t}" for i, t in enumerate(upcoming_tasks[:5])) if upcoming_tasks else ""
                    
                    _heading_prompt = f"""Create short category headings for each task below.

RULES:
- Extract the MAIN TOPIC from each task as heading (2-5 words, title case)
- DO NOT rephrase the task content - keep it exactly as given
- Keep headings simple and descriptive

Completed/In-Progress tasks:
{completed_raw}

In-Progress tasks:
{in_progress_raw}

Upcoming tasks:
{upcoming_raw}

OUTPUT FORMAT (exact JSON):
For each task, output a heading and keep the original task:
{{
  "tasks": [
    {{"heading": "2-word heading", "task": "original task exactly as given"}},
    ...
  ]
}}

Return ONLY valid JSON:"""

                    try:
                        client = _get_genai_client()
                        _head_response = client.models.generate_content(
                            model=GEMINI_MODEL,
                            contents=_heading_prompt,
                            config=_RETRY_CONFIG,
                        )
                        _head_text = _head_response.text.strip()
                        
                        # Parse JSON
                        if "{" in _head_text:
                            _head_json = json.loads(_head_text[_head_text.find("{"):_head_text.rfind("}")+1])
                            tasks_list = _head_json.get("tasks", [])
                            
                            # Build bullets with LLM-generated headings
                            completed_lines = []
                            in_progress_lines = []
                            upcoming_lines = []
                            
                            for i, task_info in enumerate(tasks_list):
                                heading = task_info.get("heading", "").strip()
                                task = task_info.get("task", "").strip()
                                if heading and task:
                                    if i < len(completed_tasks):
                                        completed_lines.append(f"• {heading}\n  {task}")
                                    elif i < len(completed_tasks) + len(in_progress_tasks):
                                        in_progress_lines.append(f"• {heading}\n  {task}")
                                    else:
                                        upcoming_lines.append(f"• {heading}\n  {task}")
                            
                            completed_bullets = "\n\n".join(completed_lines) if completed_lines else "No completed tasks"
                            in_progress_bullets = "\n\n".join(in_progress_lines) if in_progress_lines else "No in-progress tasks"
                            upcoming_bullets = "\n\n".join(upcoming_lines) if upcoming_lines else "No upcoming tasks"
                            
                            print(f"[DEBUG] Headings generated: {len(tasks_list)} tasks")
                        else:
                            raise ValueError("No JSON")
                    except Exception as _he:
                        print(f"[DEBUG] Heading LLM failed ({_he}), using direct format")
                        # Fallback - use direct format without headings
                        completed_lines = [f"• {t}" for t in completed_tasks[:5]]
                        in_progress_lines = [f"• {t}" for t in in_progress_tasks[:5]]
                        upcoming_lines = [f"• {t}" for t in upcoming_tasks[:5]]
                        
                        completed_bullets = "\n\n".join(completed_lines) if completed_lines else "No completed tasks"
                        in_progress_bullets = "\n\n".join(in_progress_lines) if in_progress_lines else "No in-progress tasks"
                        upcoming_bullets = "\n\n".join(upcoming_lines) if upcoming_lines else "No upcoming tasks"

                    except Exception as _se:
                        print(f"[DEBUG] Heading generation failed ({_se}), using fallback")
                        # Ultimate fallback
                        completed_lines = [f"• {t}" for t in completed_tasks[:5]]
                        in_progress_lines = [f"• {t}" for t in in_progress_tasks[:5]]
                        upcoming_lines = [f"• {t}" for t in upcoming_tasks[:5]]
                        
                        completed_bullets = "\n".join(completed_lines) if completed_lines else "No completed tasks"
                        in_progress_bullets = "\n".join(in_progress_lines) if in_progress_lines else "No in-progress tasks"
                        upcoming_bullets = "\n".join(upcoming_lines) if upcoming_lines else "No upcoming tasks"
                    
                    print(f"[DEBUG] Structured bullets created with headings")

                    # Update placeholders with actual data from sheet
                    placeholder_map = {
                        "{{Project Name}}": project_name or "Weekly Report",
                        "{{From Date}}": _format_indian_date(dates[0]) if dates else "",
                        "{{To Date}}": _format_indian_date(dates[-1]) if dates and len(dates) > 1 else "",
                        "{{Executive Summary}}": exec_summary,
                        "{{Key Activities Completed}}": completed_bullets,
                        "{{Key Activities In Progress}}": in_progress_bullets,
                        "{{Key Activities Upcoming}}": upcoming_bullets,
                        # Status indicators for Slide 3
                        "{{Completed}}": str(len(completed_tasks)),
                        "{{In Progress}}": str(len(in_progress_tasks)),
                        "{{Yet to Start}}": str(len(upcoming_tasks)),
                        "{{Delayed}}": "0",
                    }
                    
                    # Skip LLM rephrasing - use task data directly as-is
                    # (LLM enhancement disabled to preserve original task names)
                    print(f"[DEBUG] Using direct task data for slide content")

                    print(f"[DEBUG] Extracted placeholders: {list(placeholder_map.keys())}")
                    print(f"[DEBUG] Status counts - Completed: {len(completed_tasks)}, In Progress: {len(in_progress_tasks)}, Upcoming: {len(upcoming_tasks)}")
                    
                    # Save to state so build_slides can use it
                    tool_context.state["placeholder_map"] = placeholder_map
                    tool_context.state["project_name"] = project_name
                
                # Get timeline data for image
                # Try different sheet name variations
                sheet_variants = ["Timeline 1", "Timeline1", "Timeline", "timeline"]
                timeline_data = []
                thumbnail_link = ""
                
                # First, try to get the full spreadsheet with charts
                try:
                    full_sheet = sheets_service.spreadsheets().get(
                        spreadsheetId=sheet_id,
                        includeGridData=True
                    ).execute()
                    
                    # Look for charts in the sheets
                    for sheet in full_sheet.get('sheets', []):
                        sheet_title = sheet.get('properties', {}).get('title', '')
                        if 'timeline' in sheet_title.lower():
                            print(f"[DEBUG] Found timeline sheet: {sheet_title}")
                            # Check for charts
                            charts = sheet.get('charts', [])
                            if charts:
                                print(f"[DEBUG] Found {len(charts)} charts in Timeline sheet")
                    
                except Exception as e:
                    print(f"[DEBUG] Full sheet metadata error: {e}")
                
                # Also try values API as fallback
                for sheet_name in sheet_variants:
                    try:
                        range_str = f"'{sheet_name}'!A1:Z50"
                        print(f"[DEBUG] Trying timeline sheet values: {range_str}")
                        tl_result = sheets_service.spreadsheets().values().get(
                            spreadsheetId=sheet_id,
                            range=range_str,
                        ).execute()
                        timeline_data = tl_result.get("values", [])
                        if timeline_data:
                            print(f"[DEBUG] Found timeline sheet: {sheet_name}, rows: {len(timeline_data)}")
                            break
                    except Exception as e:
                        print(f"[DEBUG] Sheet '{sheet_name}' error: {e}")
                        continue
                    
                # Try to export Timeline sheet as image using exportLinks
                if not timeline_data:
                    try:
                        # Get the spreadsheet with metadata to find Timeline sheet ID
                        spreadsheet_metadata = drive.files().get(
                            fileId=sheet_id,
                            fields="exportLinks,name",
                            supportsAllDrives=True
                        ).execute()
                        
                        export_links = spreadsheet_metadata.get("exportLinks", {})
                        print(f"[DEBUG] Available export formats: {list(export_links.keys())}")
                        
                        # Try PDF export - we can extract images from it
                        pdf_link = export_links.get("application/pdf")
                        if pdf_link:
                            print(f"[DEBUG] PDF export available")
                            
                        # Get thumbnail - this is overall spreadsheet, not specific sheet
                        thumbnail = drive.files().get(
                            fileId=sheet_id,
                            fields="thumbnailLink",
                            supportsAllDrives=True
                        ).execute()
                        thumbnail_link = thumbnail.get("thumbnailLink", "")
                        print(f"[DEBUG] Timeline thumbnail: {thumbnail_link[:100]}...")
                        
                    except Exception as e:
                        print(f"[DEBUG] Could not get export: {e}")
                        thumbnail_link = ""
                
                if timeline_data:
                    sheet_img = _render_table_to_png(timeline_data, dpi=150)
                else:
                    # Generate Gantt chart from task data - better visualization
                    print(f"[DEBUG] timeline_data empty, checking task_data. task_data rows: {len(task_data) if task_data else 0}")
                    if task_data and len(task_data) > 1:
                        # Convert task_data rows to dict format for Gantt chart
                        task_dicts = []
                        print(f"[DEBUG] First row (header): {task_data[0]}")
                        for row in task_data[1:]:  # Skip header: ['Phases', 'Module', 'Tasks', 'Start Date', 'End Date', 'Duration', 'Status', 'Remarks']
                            if len(row) >= 7:
                                task_dicts.append({
                                    "phase": row[0] if len(row) > 0 else "",        # Phases
                                    "module": row[1] if len(row) > 1 else "",       # Module
                                    "task_detail": row[2] if len(row) > 2 else "",  # Tasks
                                    "start_date": row[3] if len(row) > 3 else "",   # Start Date
                                    "end_date": row[4] if len(row) > 4 else "",     # End Date
                                    "status": row[6] if len(row) > 6 else "",       # Status
                                })
                        print(f"[DEBUG] task_dicts created: {len(task_dicts)}")
                        if task_dicts:
                            print(f"[DEBUG] First task: {task_dicts[0]}")
                            print(f"[DEBUG] Generating Gantt chart with {len(task_dicts)} tasks")
                            sheet_img = _render_gantt_chart(task_dicts, dpi=150)
                            print(f"[DEBUG] Gantt chart generated, sheet_img: {sheet_img}")
                        else:
                            # Fallback to table
                            table_data = task_data[:15] if task_data else []
                            if table_data:
                                sheet_img = _render_table_to_png(table_data, dpi=150)
                    elif thumbnail_link:
                        # Only try thumbnail if no task data
                        try:
                            import urllib.request
                            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
                            os.close(tmp_fd)
                            import socket
                            socket.setdefaulttimeout(10)
                            urllib.request.urlretrieve(thumbnail_link, tmp_path)
                            sheet_img = Image.open(tmp_path)
                            os.remove(tmp_path)
                            print(f"[DEBUG] Using thumbnail as timeline image")
                        except Exception as e:
                            print(f"[DEBUG] Failed to download thumbnail: {e}")
                            sheet_img = None
                
                if sheet_img:
                    sheet_img_w, sheet_img_h = sheet_img.size
                    print(f"[DEBUG] Generated image from sheet: {sheet_img_w}x{sheet_img_h}")
                        
        except Exception as e:
            print(f"[DEBUG] Error reading Google Sheet: {e}")
    
    # ── 2. Duplicate template ──────────────────────────────────────────────
    date_str   = datetime.today().strftime("%d %b %Y")
    copy_title = f"{project_name} - Status Report - {date_str}"

    try:
        copy_body = {"name": copy_title}
        if PARENT_FOLDER_ID:
            copy_body["parents"] = [PARENT_FOLDER_ID]

        copied = drive.files().copy(
            fileId=template_id,
            body=copy_body,
            fields="id,webViewLink",
            supportsAllDrives=True,
        ).execute()
    except HttpError as exc:
        return {"success": False, "slides_url": "", "slides_id": "",
                "error": f"Template copy failed: {exc}"}

    slides_id  = copied.get("id", "")
    slides_url = copied.get(
        "webViewLink",
        f"https://docs.google.com/presentation/d/{slides_id}/edit"
    )

    # ── 2a. Fill Slide 2 text boxes by Alt Text Title ────────────────
    # More reliable than find-and-replace for nested text boxes
    # Map Alt Text titles (set in template) to placeholder_map keys
    alt_text_map = {
        "{{EXECUTIVE_SUMMARY}}": placeholder_map.get("{{Executive Summary}}", ""),
        "{{COMPLETED_ACTIVITIES}}": placeholder_map.get("{{Key Activities Completed}}", "") + "\n\n" + placeholder_map.get("{{Key Activities In Progress}}", ""),
        "{{UPCOMING_ACTIVITIES}}": placeholder_map.get("{{Key Activities Upcoming}}", ""),
    }
    normalized_alt_text_map = {}
    for tag, content in alt_text_map.items():
        normalized_content = str(content or "")
        if tag in ("{{COMPLETED_ACTIVITIES}}", "{{UPCOMING_ACTIVITIES}}"):
            cleaned_lines = []
            for line in normalized_content.splitlines():
                stripped_line = line.lstrip()
                stripped_line = re.sub(r"^[•○◦●▪\-]+\s*", "", stripped_line)
                cleaned_lines.append(stripped_line if line.strip() else "")
            normalized_content = "\n".join(cleaned_lines)
        normalized_alt_text_map[tag] = normalized_content
    formatted_alt_text_map = {}
    for tag, content in normalized_alt_text_map.items():
        if tag not in ("{{COMPLETED_ACTIVITIES}}", "{{UPCOMING_ACTIVITIES}}"):
            formatted_alt_text_map[tag] = str(content or "").strip()
            continue
        # Build native nested bullets by using plain headings (level 1)
        # and tab-indented detail lines (level 2) before createParagraphBullets.
        formatted_lines = []
        expect_heading = True
        for raw_line in str(content or "").splitlines():
            line = raw_line.strip()
            if not line:
                expect_heading = True
                continue
            if expect_heading:
                formatted_lines.append(line)
                expect_heading = False
            else:
                formatted_lines.append(f"\t{line}")
                expect_heading = True
        formatted_alt_text_map[tag] = "\n".join(formatted_lines)
    print(f"[DEBUG] Alt text map with data: {[(k, len(v)) for k,v in alt_text_map.items()]}")
    
    # Fetch all slides to find shapes by their title (Alt Text)
    try:
        prs = slides.presentations().get(presentationId=slides_id).execute()
        all_slides = prs.get("slides", [])
        
        # Debug: log all element titles/descriptions found in slides
        all_titles = []
        all_descriptions = []
        for slide in all_slides:
            for elem in slide.get("pageElements", []):
                title = (elem.get("title", "") or "").strip()
                desc = (elem.get("description", "") or "").strip()
                if title:
                    all_titles.append(title)
                if desc:
                    all_descriptions.append(desc)
        print(f"[DEBUG] All element titles found: {all_titles}")
        print(f"[DEBUG] All element descriptions found: {all_descriptions}")
        
        # Collect update requests for text boxes
        text_update_requests = []
        style_requests = []
        bullet_only_requests = []
        heading_bold_requests = []
        activity_object_ids = set()
        
        for tag, content in formatted_alt_text_map.items():
            if not content:
                continue
            for slide in all_slides:
                for elem in slide.get("pageElements", []):
                    elem_title = (elem.get("title", "") or "").strip()
                    elem_desc = (elem.get("description", "") or "").strip()
                    elem_id = elem.get("objectId", "")
                    
                    normalized_tag = re.sub(r"\s+", "", tag).upper()
                    normalized_title = re.sub(r"\s+", "", elem_title).upper()
                    normalized_desc = re.sub(r"\s+", "", elem_desc).upper()
                    tag_matched = (
                        normalized_title == normalized_tag
                        or normalized_desc == normalized_tag
                        or (normalized_tag in normalized_title if normalized_title else False)
                        or (normalized_tag in normalized_desc if normalized_desc else False)
                    )
                    if tag_matched:
                        if "shape" in elem:
                            shape_text = elem.get("shape", {}).get("text", {})
                            has_text = any(
                                te.get("textRun", {}).get("content", "").strip()
                                for te in shape_text.get("textElements", [])
                            )
                            if has_text:
                                text_update_requests.append({
                                    "deleteText": {
                                        "objectId": elem_id,
                                        "textRange": {"type": "ALL"},
                                    }
                                })
                            text_update_requests.append({
                                "insertText": {
                                    "objectId": elem_id,
                                    "insertionIndex": 0,
                                    "text": content
                                }
                            })

                            # Apply DM Sans to inserted content for exact visual parity.
                            # Executive summary uses 21pt, others use 16pt
                            font_size = 21 if tag == "{{EXECUTIVE_SUMMARY}}" else 16
                            style_requests.append({
                                "updateTextStyle": {
                                    "objectId": elem_id,
                                    "textRange": {"type": "ALL"},
                                    "style": {
                                        "fontFamily": "DM Sans",
                                        "fontSize": {
                                            "magnitude": font_size,
                                            "unit": "PT",
                                        },
                                        "bold": False,
                                    },
                                    "fields": "fontFamily,fontSize,bold",
                                }
                            })

                            # Prepare official Slides bullets (not manual symbols) for activity boxes.
                            if tag in ("{{COMPLETED_ACTIVITIES}}", "{{UPCOMING_ACTIVITIES}}"):
                                activity_object_ids.add(elem_id)
                                bullet_only_requests.append({
                                    "createParagraphBullets": {
                                        "objectId": elem_id,
                                        "textRange": {"type": "ALL"},
                                        "bulletPreset": "BULLET_DISC_CIRCLE_SQUARE",
                                    }
                                })
                            print(f"[DEBUG] Alt Text fill: '{tag}' -> '{content[:30]}...'")

        # Execute text insert/delete first so content always lands.
        if text_update_requests:
            try:
                slides.presentations().batchUpdate(
                    presentationId=slides_id,
                    body={"requests": text_update_requests}
                ).execute()
                print(f"[DEBUG] Alt Text updates applied: {len(text_update_requests)} fields")
            except Exception as batch_exc:
                print(f"[WARNING] Text insert batch failed: {batch_exc}")

        # Then apply base font style.
        if style_requests:
            try:
                slides.presentations().batchUpdate(
                    presentationId=slides_id,
                    body={"requests": style_requests}
                ).execute()
                print(f"[DEBUG] Text style applied: {len(style_requests)} fields")
            except Exception as style_exc:
                print(f"[WARNING] Base style batch failed: {style_exc}")

        # Apply official Slides bullets in a separate batch to avoid range conflicts.
        if bullet_only_requests:
            try:
                slides.presentations().batchUpdate(
                    presentationId=slides_id,
                    body={"requests": bullet_only_requests}
                ).execute()
                print(f"[DEBUG] Official bullet formatting applied: {len(bullet_only_requests)} fields")
            except Exception as bullet_exc:
                print(f"[WARNING] Bullet formatting batch failed: {bullet_exc}")

        # Finally bold heading lines only (after bullets exist and tabs are normalized by API).
        if activity_object_ids:
            print(f"[DEBUG] Activity object IDs to bold: {activity_object_ids}")
            try:
                latest_prs = slides.presentations().get(presentationId=slides_id).execute()
                for slide in latest_prs.get("slides", []):
                    for elem in slide.get("pageElements", []):
                        obj_id = elem.get("objectId", "")
                        if obj_id not in activity_object_ids:
                            continue
                        text_content = (
                            elem.get("shape", {})
                                .get("text", {})
                                .get("text", "")
                        )
                        text_elements = (
                            elem.get("shape", {})
                                .get("text", {})
                                .get("textElements", [])
                        )
                        print(f"[DEBUG] Shape {obj_id} has {len(text_elements)} text elements, content length: {len(text_content)}")
                        print(f"[DEBUG] Content preview: {text_content[:100]}...")
                        
                        for idx, te in enumerate(text_elements):
                            pm = te.get("paragraphMarker", {})
                            bullet = pm.get("bullet", {})
                            if not bullet:
                                continue
                            nesting = bullet.get("nestingLevel", "0")
                            if int(nesting) != 0:
                                continue
                            start_idx = te.get("startIndex")
                            end_idx = te.get("endIndex")
                            if start_idx is None:
                                start_idx = 0
                            if not isinstance(start_idx, int) or not isinstance(end_idx, int):
                                print(f"[DEBUG] Skipping element {idx}: start={start_idx} (type={type(start_idx)}), end={end_idx} (type={type(end_idx)})")
                                continue
                            if end_idx - 1 <= start_idx:
                                continue
                            heading_bold_requests.append({
                                "updateTextStyle": {
                                    "objectId": obj_id,
                                    "textRange": {
                                        "type": "FIXED_RANGE",
                                        "startIndex": start_idx,
                                        "endIndex": end_idx - 1,
                                    },
                                    "style": {"bold": True},
                                    "fields": "bold",
                                }
                            })
                            print(f"[DEBUG] Added bold request for element {idx}")
            except Exception as heading_scan_exc:
                print(f"[WARNING] Heading scan failed: {heading_scan_exc}")

        if heading_bold_requests:
            try:
                slides.presentations().batchUpdate(
                    presentationId=slides_id,
                    body={"requests": heading_bold_requests}
                ).execute()
                print(f"[DEBUG] Heading bold applied: {len(heading_bold_requests)} ranges")
            except Exception as bold_exc:
                print(f"[WARNING] Heading bold batch failed: {bold_exc}")
    except Exception as e:
        print(f"[DEBUG] Alt Text fill error (non-fatal): {e}")
    
    # ── 2b. Find and Replace (exact match) ─
    # Using matchCase: true for exact matching as per template
    replace_requests = []
    for tag, replacement in placeholder_map.items():
        # Use exact tag with braces, matchCase: true
        replace_requests.append({
            "replaceAllText": {
                "containsText": {"text": tag, "matchCase": True},
                "replaceText":  replacement,
            }
        })

    if replace_requests:
        try:
            result = slides.presentations().batchUpdate(
                presentationId=slides_id,
                body={"requests": replace_requests},
            ).execute()
            
            # Log replacement counts
            for r in result.get("replies", []):
                if "replaceAllText" in r:
                    count = r["replaceAllText"].get("occurrencesChanged", 0)
                    print(f"[DEBUG] Text replacement: {count} occurrences")
        except HttpError as exc:
            print(f"[WARNING] Text replacement error: {exc}")

    # ── 3. Timeline image injection ────────────────────────────────────────
    # Use image from Google Sheet if available, otherwise use uploaded image
    final_img = sheet_img if sheet_img else None
    final_w = sheet_img_w if sheet_img else img_w
    final_h = sheet_img_h if sheet_img else img_h
    
    if final_img and final_w > 0 and final_h > 0:
        # Upload the generated image to Drive
        try:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix="_timeline.png")
            os.close(tmp_fd)
            # Save with lower DPI for faster upload
            final_img.save(tmp_path, "PNG", dpi=(96, 96))
            print(f"[DEBUG] Saved image to temp: {tmp_path}")
            
            file_meta = {
                "name": f"timeline_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                "mimeType": "image/png",
            }
            if PARENT_FOLDER_ID:
                file_meta["parents"] = [PARENT_FOLDER_ID]
            
            # Upload with timeout
            media = MediaFileUpload(tmp_path, mimetype="image/png", resumable=True, chunksize=1024*1024)
            uploaded = drive.files().create(
                body=file_meta, media_body=media,
                fields="id,webContentLink",
                supportsAllDrives=True
            ).execute()
            
            drive_image_id = uploaded.get("id", "")
            final_w, final_h = final_img.size
            print(f"[DEBUG] Uploaded sheet image to Drive: {drive_image_id}")
            
            # Make readable
            drive.permissions().create(
                fileId=drive_image_id,
                body={"type": "anyone", "role": "reader"},
                supportsAllDrives=True,
            ).execute()
            
            os.remove(tmp_path)
            
        except Exception as e:
            print(f"[WARNING] Failed to upload sheet image: {e}")
            drive_image_id = ""
            final_img = None
    else:
        print(f"[DEBUG] No valid image to upload - using fallback drive_image_id")
    
    print(f"[DEBUG] Timeline image injection: drive_image_id={drive_image_id}, img_w={final_w}, img_h={final_h}")
    if drive_image_id and final_img:
        try:
            _inject_timeline_image(
                slides, slides_id, drive_image_id, final_w, final_h
            )
            print(f"[DEBUG] Timeline image injected successfully")
        except Exception as exc:
            print(f"[WARNING] Timeline image injection error (non-fatal): {exc}")

    state["slides_id"]  = slides_id
    state["slides_url"] = slides_url

    return {"success": True, "slides_url": slides_url,
            "slides_id": slides_id, "error": ""}


def _inject_timeline_image(
    slides_svc,
    presentation_id: str,
    drive_image_id:  str,
    img_w_px:  int,
    img_h_px:  int,
):
    """
    Find the shape tagged {{TIMELINE_SCREENSHOT}} in the presentation,
    delete it, and insert the PNG image smart-scaled to fit the slide.

    Smart scaling rules (matching the spec):
      - If height-constrained (tall image): fit to slide height, scale width.
      - If width-constrained  (wide image): fit to slide width (10 in), scale height.
      - Lock aspect ratio — never stretch.
    """
    # ── a. Fetch presentation to find the placeholder shape ───────────────
    prs       = slides_svc.presentations().get(
        presentationId=presentation_id
    ).execute()
    slides    = prs.get("slides", [])

    placeholder_element_id = None
    placeholder_slide_id   = None
    placeholder_pos        = None   # {translateX, translateY, scaleX, scaleY}

    # Search by shape title/name OR by text content containing placeholder
    for slide in slides:
        for element in slide.get("pageElements", []):
            # Check shape name/title first
            shape_title = element.get("title", "")
            shape_description = element.get("description", "")
            
            # Check if shape name contains "timeline" or "gantt"
            if "timeline" in shape_title.lower() or "gantt" in shape_title.lower():
                placeholder_element_id = element["objectId"]
                placeholder_slide_id = slide["objectId"]
                tf = element.get("transform", {})
                size = element.get("size", {})
                placeholder_pos = {
                    "translateX": tf.get("translateX", 0),
                    "translateY": tf.get("translateY", 0),
                    "scaleX":     tf.get("scaleX", 1),
                    "scaleY":     tf.get("scaleY", 1),
                    "width":      size.get("width",  {}).get("magnitude", SLIDE_W_PX * 9144),
                    "height":     size.get("height", {}).get("magnitude", SLIDE_H_PX * 9144),
                }
                print(f"[DEBUG] Found placeholder by name: {shape_title}")
                break
            
            # Also check text content
            shape = element.get("shape", {})
            text  = shape.get("text", {})
            for te in text.get("textElements", []):
                content = te.get("textRun", {}).get("content", "")
                if "{{TIMELINE_SCREENSHOT}}" in content:
                    placeholder_element_id = element["objectId"]
                    placeholder_slide_id   = slide["objectId"]
                    tf = element.get("transform", {})
                    size = element.get("size", {})
                    placeholder_pos = {
                        "translateX": tf.get("translateX", 0),
                        "translateY": tf.get("translateY", 0),
                        "scaleX":     tf.get("scaleX", 1),
                        "scaleY":     tf.get("scaleY", 1),
                        "width":      size.get("width",  {}).get("magnitude", SLIDE_W_PX * 9144),
                        "height":     size.get("height", {}).get("magnitude", SLIDE_H_PX * 9144),
                    }
                    break
            if placeholder_element_id:
                break
        if placeholder_element_id:
            break

    # ── b. Smart-scale the image to fit the placeholder (or full slide) ───
    # Google Slides uses EMU (English Metric Units): 1 inch = 914400 EMU
    EMU_PER_INCH = 914400

    page_size = prs.get("pageSize", {})
    page_w_emu = page_size.get("width", {}).get("magnitude")
    page_h_emu = page_size.get("height", {}).get("magnitude")
    if not page_w_emu or not page_h_emu:
        page_w_emu = int(SLIDE_W_IN * EMU_PER_INCH)
        page_h_emu = int(SLIDE_H_IN * EMU_PER_INCH)

    if placeholder_pos:
        # Use the placeholder's own bounding box (account for scale)
        scale_x = placeholder_pos.get("scaleX", 1) or 1
        scale_y = placeholder_pos.get("scaleY", 1) or 1
        avail_w_emu = int(placeholder_pos["width"] * scale_x)
        avail_h_emu = int(placeholder_pos["height"] * scale_y)
        left_emu    = int(placeholder_pos["translateX"])
        top_emu     = int(placeholder_pos["translateY"])
        print(f"[DEBUG] Using placeholder bounds for timeline image")
    else:
        # No placeholder found — use a centered area on slide 3
        avail_w_emu = int(page_w_emu * 0.95)
        avail_h_emu = int(page_h_emu * 0.85)
        left_emu    = int((page_w_emu - avail_w_emu) // 2)  # Center horizontally
        top_emu     = int((page_h_emu - avail_h_emu) // 2)   # Center vertically
        placeholder_slide_id = slides[2]["objectId"] if len(slides) > 2 else (slides[0]["objectId"] if slides else None)
        print(f"[DEBUG] Using slide 3 for timeline image: {placeholder_slide_id}")

    if not placeholder_slide_id:
        raise ValueError("No slides found in the presentation.")

    print(f"[DEBUG] Image original: {img_w_px}x{img_h_px}px")
    print(f"[DEBUG] Available space: {avail_w_emu//914000}x{avail_h_emu//914000} inches")
    
    # Scale image proportionally to fit within available space while maintaining aspect ratio
    img_aspect = img_w_px / img_h_px if img_h_px > 0 else 1.0
    avail_aspect = avail_w_emu / avail_h_emu if avail_h_emu > 0 else 1.0
    
    if img_aspect > avail_aspect:
        # Image is wider than available space - fit to width
        fit_w_emu = avail_w_emu
        fit_h_emu = int(fit_w_emu / img_aspect)
    else:
        # Image is taller than available space - fit to height
        fit_h_emu = avail_h_emu
        fit_w_emu = int(fit_h_emu * img_aspect)
    
    # Centre within the available area
    center_left = left_emu + (avail_w_emu - fit_w_emu) // 2
    center_top  = top_emu  + (avail_h_emu - fit_h_emu) // 2

    # ── c. Build batchUpdate requests ────────────────────────────────────
    requests = []

    # Insert the PNG image from Drive
    new_image_id = f"timeline_img_{datetime.now().strftime('%H%M%S%f')}"
    image_url    = f"https://drive.google.com/uc?export=download&id={drive_image_id}"
    
    print(f"[DEBUG] Image URL: {image_url}")
    print(f"[DEBUG] Placing on slide: {placeholder_slide_id}")
    print(f"[DEBUG] Image size: {fit_w_emu}x{fit_h_emu} EMU")

    requests.append({
        "createImage": {
            "objectId":    new_image_id,
            "url":         image_url,
            "elementProperties": {
                "pageObjectId": placeholder_slide_id,
                "size": {
                    "width":  {"magnitude": fit_w_emu,    "unit": "EMU"},
                    "height": {"magnitude": fit_h_emu,    "unit": "EMU"},
                },
                "transform": {
                    "scaleX":     1,
                    "scaleY":     1,
                    "translateX": center_left,
                    "translateY": center_top,
                    "unit":       "EMU",
                },
            },
        }
    })

    result = slides_svc.presentations().batchUpdate(
        presentationId=presentation_id,
        body={"requests": requests},
    ).execute()
    
    print(f"[DEBUG] Image create result: {result}")

    # Delete the placeholder shape after image insertion
    if placeholder_element_id:
        try:
            slides_svc.presentations().batchUpdate(
                presentationId=presentation_id,
                body={"requests": [{"deleteObject": {"objectId": placeholder_element_id}}]},
            ).execute()
        except Exception as delete_exc:
            print(f"[WARNING] Placeholder delete failed: {delete_exc}")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 3 — TOOL 4: CLEANUP
# ══════════════════════════════════════════════════════════════════════════════

def cleanup_temp_assets(*, tool_context: ToolContext) -> dict:
    """
    Delete the temporary timeline PNG from Google Drive.

    Reads session state:
        drive_image_id (str)

    Returns:
        success (bool), deleted (bool), error (str)
    """
    state          = tool_context.state
    drive_image_id = state.get("drive_image_id", "")

    deleted = False
    drive_error = ""

    if drive_image_id:
        try:
            drive, _ = _build_services()
            drive.files().delete(
                fileId=drive_image_id,
                supportsAllDrives=True,
            ).execute()
            state["drive_image_id"] = ""   # clear from state
            deleted = True
        except HttpError as exc:
            drive_error = f"Drive delete failed: {exc}"

    # Clean up any temp .xlsx created from legacy .xls conversion
    converted_path = state.get("converted_excel_path", "")
    converted_dir = state.get("converted_excel_dir", "")
    if converted_path and os.path.exists(converted_path):
        try:
            os.remove(converted_path)
        except OSError:
            pass
    if converted_dir and os.path.isdir(converted_dir):
        try:
            os.rmdir(converted_dir)
        except OSError:
            pass

    state["converted_excel_path"] = ""
    state["converted_excel_dir"] = ""

    if drive_image_id and drive_error:
        return {"success": False, "deleted": deleted, "error": drive_error}

    if not drive_image_id:
        return {"success": True, "deleted": False,
                "error": "No drive_image_id in state — nothing to delete."}

    return {"success": True, "deleted": deleted, "error": ""}


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 5 — STATE READER
# ══════════════════════════════════════════════════════════════════════════════

def get_slides_state(*, tool_context: ToolContext) -> dict:
    """
    Return a snapshot of all slides pipeline state for the summary agent.

    Returns:
        project_name (str), timeline_range (str), placeholder_count (int),
        image_dimensions (str), slides_url (str), slides_id (str), success (bool)
    """
    state = tool_context.state
    pm    = _safe_json(state.get("placeholder_map")) or {}
    w     = state.get("image_width_px",  0)
    h     = state.get("image_height_px", 0)
    return {
        "success":          True,
        "project_name":     state.get("project_name",    ""),
        "timeline_range":   state.get("timeline_range",  ""),
        "timeline_error":   state.get("timeline_error",  ""),
        "placeholder_count": len(pm),
        "image_dimensions": f"{w} × {h} px" if w and h else "N/A",
        "slides_url":       state.get("slides_url", ""),
        "slides_id":        state.get("slides_id",  ""),
    }


# ══════════════════════════════════════════════════════════════════════════════
# AGENTS
# ══════════════════════════════════════════════════════════════════════════════

excel_reader_agent = LlmAgent(
    name="ExcelReaderAgent",
    model=GEMINI_MODEL,
    description=(
        "Parses the Excel weekly report to extract project metadata, "
        "placeholder values, and the dynamic Timeline sheet range."
    ),
    before_model_callback=_strip_unsupported_excel_attachments,
    instruction="""
You parse an Excel report file. Make exactly ONE tool call.

Call: extract_excel_data(excel_file_path="[path from user message]")

The path comes from the user's message - look for a local .xlsx/.xlsm/.xltx/.xltm path (or a file:// URI). If the user provides a legacy .xls file, conversion will be attempted.

After the call reply with plain text:
- project name found
- number of placeholders extracted
- timeline range detected (e.g. A1:M32)
- any timeline sheet errors
- confirmation that screenshot agent can proceed
""",
    tools=[extract_excel_data],
    generate_content_config=_RETRY_CONFIG,
)


screenshot_agent = LlmAgent(
    name="ScreenshotAgent",
    model=GEMINI_MODEL,
    description=(
        "Renders the Timeline sheet range to a high-resolution PNG "
        "and uploads it to Google Drive."
    ),
    before_model_callback=_strip_unsupported_excel_attachments,
    instruction="""
Render the timeline screenshot. Make exactly ONE tool call.

Call: capture_timeline_screenshot()  - no arguments required.

If success=False, report the exact error and stop.
After success reply with the width_px, height_px and drive_image_id values from the tool response.
""",
    tools=[capture_timeline_screenshot],
    generate_content_config=_RETRY_CONFIG,
)


slides_builder_agent = LlmAgent(
    name="SlidesBuilderAgent",
    model=GEMINI_MODEL,
    description=(
        "Duplicates the Google Slides template, injects all placeholder text, "
        "and embeds the timeline screenshot with smart aspect-ratio scaling."
    ),
    before_model_callback=_strip_unsupported_excel_attachments,
    instruction="""
Build the Google Slides report. The tool will:
1. Replace placeholders case-insensitively (e.g., tags like PROJECT_NAME match Project Name)
2. Inject the timeline screenshot

Make exactly ONE tool call:

Call: build_slides_report()  - no arguments required.

If success=False, report the exact error and stop.
After success reply with the slides_url from the tool response.
""",
    tools=[build_slides_report],
    generate_content_config=_RETRY_CONFIG,
)


cleanup_agent = LlmAgent(
    name="CleanupAgent",
    model=GEMINI_MODEL,
    description=(
        "Deletes the temporary Drive image and produces the final "
        "summary with the Google Slides URL."
    ),
    before_model_callback=_strip_unsupported_excel_attachments,
    instruction="""
Make exactly 2 tool calls in order:
1. cleanup_temp_assets()  — no arguments
2. get_slides_state()     — no arguments

The get_slides_state() tool returns: project_name, timeline_range, placeholder_count, image_dimensions, slides_url

After both complete, write a final markdown summary with the actual values from the tool responses.

## Report Generated Successfully

Project: [use project_name from get_slides_state]
Timeline Range Captured: [use timeline_range]
Placeholders Injected: [use placeholder_count]
Image Dimensions: [use image_dimensions]

### Google Slides Report
Link: [use slides_url from get_slides_state]

If any step failed, report the error clearly.
""",
    tools=[cleanup_temp_assets, get_slides_state],
    generate_content_config=_RETRY_CONFIG,
)


# ══════════════════════════════════════════════════════════════════════════════
# ROOT AGENT
# ══════════════════════════════════════════════════════════════════════════════

root_agent = SequentialAgent(
    name="SlidesReportOrchestrator",
    description=(
        "Phase 2 pipeline: reads an Excel weekly report, dynamically detects "
        "the Timeline sheet range, renders it as a PNG, duplicates the Google "
        "Slides template, injects all placeholder text and the timeline image "
        "(smart-scaled), cleans up Drive, and returns the final Slides URL."
    ),
    sub_agents=[
        excel_reader_agent,
        screenshot_agent,
        slides_builder_agent,
        cleanup_agent,
    ],
)

__all__ = ["root_agent"]


# ══════════════════════════════════════════════════════════════════════════════
# STANDALONE TEST  (python slides_agent.py)
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import asyncio
    import sys
    from google.adk.runners import Runner
    from google.adk.sessions import InMemorySessionService

    excel_path = sys.argv[1] if len(sys.argv) > 1 else "./outputs/sample_report.xlsx"

    async def run():
        svc     = InMemorySessionService()
        session = await svc.create_session(app_name="slides_report", user_id="u1")
        runner  = Runner(agent=root_agent, app_name="slides_report",
                         session_service=svc)
        msg     = genai_types.Content(
            role="user",
            parts=[genai_types.Part(
                text=f"Generate the Google Slides report from this Excel file: {excel_path}"
            )]
        )
        print("\n" + "=" * 70)
        print(f"  Slides Report Agent  |  Model: {GEMINI_MODEL}")
        print("=" * 70 + "\n")
        async for event in runner.run_async(
                user_id="u1", session_id=session.id, new_message=msg):
            author = getattr(event, "author", "System")
            if event.is_final_response():
                if event.content and event.content.parts:
                    for part in event.content.parts:
                        if getattr(part, "text", ""):
                            print(f"\n{'─'*60}\n  FINAL [{author}]\n{'─'*60}")
                            print(part.text)
            else:
                if event.content and event.content.parts:
                    for part in event.content.parts:
                        txt = (getattr(part, "text", None) or "").strip()
                        if txt:
                            print(f"[{author}] {txt[:140]}")

    asyncio.run(run())
